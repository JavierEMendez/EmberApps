"""
excel_export.py — fills the Ember template Excel with project inputs and returns bytes.
"""
import io, copy, os
import openpyxl
from openpyxl.styles import PatternFill

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "Ember_Template.xlsx")

def _s(v):
    """Safe string."""
    return v if v is not None else ""

def _n(v, default=None):
    """Safe number."""
    try:
        return float(v) if v not in (None, "", "—") else default
    except (ValueError, TypeError):
        return default

def _pct(v):
    """Convert percentage display (e.g. 17 → 0.17)."""
    n = _n(v)
    return n / 100 if n is not None else None

def write_cell(ws, cell_addr, value):
    """Write a value to a cell, leaving formula cells alone if value is None."""
    if value is None:
        return
    ws[cell_addr] = value


def export_excel(inputs: dict) -> bytes:
    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # ── TRACT INPUTS ─────────────────────────────────────────────────
    ti = wb["Tract Inputs"]
    write_cell(ti, "B5",  _s(inputs.get("project_name")))
    write_cell(ti, "B6",  _s(inputs.get("address")))
    write_cell(ti, "B7",  _n(inputs.get("gross_acreage")))
    write_cell(ti, "B8",  _n(inputs.get("land_escalator")))
    write_cell(ti, "B9",  _n(inputs.get("purchase_price_per_acre")))
    write_cell(ti, "B11", _n(inputs.get("closing_costs_pct")))
    closing_date = inputs.get("closing_date")
    if closing_date:
        write_cell(ti, "B14", closing_date)

    # Plants (rows 19–26, col B = type)
    plants = inputs.get("plants", [])
    plant_rows = [19, 20, 21, 22, 23, 24, 25, 26]
    for i, row in enumerate(plant_rows):
        p = plants[i] if i < len(plants) else {}
        write_cell(ti, f"B{row}", _s(p.get("type", "None")))

    # Detention
    write_cell(ti, "B31", _n(inputs.get("det_storage_rate")))
    write_cell(ti, "B33", _n(inputs.get("det_depth")))
    write_cell(ti, "B34", _n(inputs.get("det_num_projects")))

    # Amenities (rows 42–47, col B = type, col C = acres)
    amenities = inputs.get("amenities", [])
    amenity_rows = [42, 43, 44, 45, 46, 47]
    for i, row in enumerate(amenity_rows):
        a = amenities[i] if i < len(amenities) else {}
        write_cell(ti, f"B{row}", _s(a.get("type", "None")))
        write_cell(ti, f"C{row}", _n(a.get("acres")))

    # Parks / drill sites
    write_cell(ti, "B51", _n(inputs.get("parks_pct")))
    write_cell(ti, "C52", _n(inputs.get("drill_site_acres")))

    # Other net-outs (rows 56–61, col A = desc, col B = acres)
    other_no = inputs.get("other_netouts", [])
    other_rows = [56, 57, 58, 59, 60, 61]
    for i, row in enumerate(other_rows):
        o = other_no[i] if i < len(other_no) else {}
        write_cell(ti, f"A{row}", _s(o.get("description")))
        write_cell(ti, f"B{row}", _n(o.get("acres")))

    # Roads (rows 65–70)
    roads = inputs.get("roads", [])
    road_rows = [65, 66, 67, 68, 69, 70]
    for i, row in enumerate(road_rows):
        r = roads[i] if i < len(roads) else {}
        write_cell(ti, f"B{row}", _s(r.get("type")))
        write_cell(ti, f"C{row}", _n(r.get("linear_feet")))
        write_cell(ti, f"D{row}", _n(r.get("width")))
        write_cell(ti, f"E{row}", _n(r.get("road_setback")))
        write_cell(ti, f"F{row}", _n(r.get("landscaping_setback")))

    # Pod acres
    write_cell(ti, "B76", _n(inputs.get("commercial_pod_acres")))
    write_cell(ti, "B77", _n(inputs.get("residential_pod_acres")))

    # ── COST INPUTS ───────────────────────────────────────────────────
    ci = wb["Cost Inputs"]

    # Global settings
    write_cell(ci, "B5",  _n(inputs.get("default_other_pct")))
    write_cell(ci, "B6",  _n(inputs.get("sectional_other_pct")))
    write_cell(ci, "B7",  _n(inputs.get("landscaping_other_pct")))
    write_cell(ci, "B8",  _n(inputs.get("contingency")))
    write_cell(ci, "B9",  _n(inputs.get("site_work_pct")))
    write_cell(ci, "B10", _n(inputs.get("fenced_pct")))
    write_cell(ci, "B11", _n(inputs.get("cost_per_mailbox")))
    write_cell(ci, "B12", _n(inputs.get("cost_per_streetlight")))
    write_cell(ci, "B13", _n(inputs.get("default_start_month")))

    # Land takedowns — period and pct (rows 17-18, cols B/C/D = take1/2/3)
    takedowns = inputs.get("takedowns") or inputs.get("land_takedowns") or []
    td_cols = ["B", "C", "D"]
    for i, col in enumerate(td_cols):
        td = takedowns[i] if i < len(takedowns) else {}
        write_cell(ci, f"{col}17", _n(td.get("period")))
        write_cell(ci, f"{col}18", _n(td.get("pct")))

    # Plant costs (rows 25–32)
    plant_costs = inputs.get("plant_costs", [])
    pc_rows = [25, 26, 27, 28, 29, 30, 31, 32]
    for i, row in enumerate(pc_rows):
        pc = plant_costs[i] if i < len(plant_costs) else {}
        write_cell(ci, f"D{row}", _n(pc.get("base_cost")))
        write_cell(ci, f"E{row}", _n(pc.get("other_pct")))
        write_cell(ci, f"G{row}", _n(pc.get("start_month")))
        write_cell(ci, f"J{row}", _n(pc.get("ph2_base_cost")))
        write_cell(ci, f"K{row}", _n(pc.get("ph2_other_pct")))
        write_cell(ci, f"M{row}", _n(pc.get("ph2_start_month")))

    # Amenity costs (rows 36–41)
    amenity_costs = inputs.get("amenity_costs", [])
    ac_rows = [36, 37, 38, 39, 40, 41]
    for i, row in enumerate(ac_rows):
        ac = amenity_costs[i] if i < len(amenity_costs) else {}
        write_cell(ci, f"D{row}", _n(ac.get("base_cost")))
        write_cell(ci, f"E{row}", _n(ac.get("other_pct")))
        write_cell(ci, f"G{row}", _n(ac.get("start_month")))

    # Detention costs (rows 45–50)
    det_costs = inputs.get("det_costs", [])
    det_rows = [45, 46, 47, 48, 49, 50]
    for i, row in enumerate(det_rows):
        dc = det_costs[i] if i < len(det_costs) else {}
        write_cell(ci, f"D{row}", _n(dc.get("other_pct")))
        write_cell(ci, f"F{row}", _n(dc.get("start_month")))
        write_cell(ci, f"I{row}", _n(dc.get("landscaping_per_foot")))

    # Other costs (rows 54–59)
    other_costs = inputs.get("other_costs", [])
    oc_rows = [54, 55, 56, 57, 58, 59]
    for i, row in enumerate(oc_rows):
        oc = other_costs[i] if i < len(other_costs) else {}
        write_cell(ci, f"D{row}", _n(oc.get("base_cost")))
        write_cell(ci, f"E{row}", _n(oc.get("other_pct")))
        write_cell(ci, f"G{row}", _n(oc.get("start_month")))
        write_cell(ci, f"H{row}", _n(oc.get("duration")))

    # Road costs (rows 63–68)
    road_costs = inputs.get("road_costs", [])
    rc_rows = [63, 64, 65, 66, 67, 68]
    for i, row in enumerate(rc_rows):
        rc = road_costs[i] if i < len(road_costs) else {}
        write_cell(ci, f"D{row}", _n(rc.get("wsd_per_lf")))
        write_cell(ci, f"E{row}", _n(rc.get("paving_per_lf")))
        write_cell(ci, f"G{row}", _n(rc.get("other_pct")))
        write_cell(ci, f"I{row}", _n(rc.get("start_month")))
        write_cell(ci, f"L{row}", _n(rc.get("landscaping_per_sf")))
        write_cell(ci, f"N{row}", _n(rc.get("light_spacing")))

    # Lot size mix (rows 72–87, 16 lot sizes 25-100 FF)
    lot_sizes = inputs.get("lot_sizes", [])
    ls_rows = [72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 84, 85, 86, 87]
    for i, row in enumerate(ls_rows):
        ls = lot_sizes[i] if i < len(lot_sizes) else {}
        on_val = ls.get("on", 0)
        write_cell(ci, f"B{row}", 1 if on_val else 0)
        write_cell(ci, f"C{row}", _n(ls.get("yield_per_ac")))
        write_cell(ci, f"D{row}", _n(ls.get("pace")))
        write_cell(ci, f"H{row}", _n(ls.get("home_price")))
        write_cell(ci, f"I{row}", _n(ls.get("wsd_per_ff")))
        write_cell(ci, f"J{row}", _n(ls.get("paving_per_ff")))
        write_cell(ci, f"L{row}", _n(ls.get("dev_start_month")))
        write_cell(ci, f"M{row}", _n(ls.get("landscaping_per_lot")))
        write_cell(ci, f"N{row}", _n(ls.get("urd_per_lot")))
        write_cell(ci, f"O{row}", _n(ls.get("lots_per_streetlight")))
        write_cell(ci, f"P{row}", _n(ls.get("fence_cost_per_ff")))

    # Overhead costs
    write_cell(ci, "B95",  _n(inputs.get("prof_svc_pct")))
    write_cell(ci, "B99",  _n(inputs.get("dmf_pct")))
    write_cell(ci, "C103", _n(inputs.get("personnel_monthly")))
    write_cell(ci, "C104", _n(inputs.get("marketing_personnel_monthly")))
    write_cell(ci, "C108", _n(inputs.get("legal_monthly")))
    write_cell(ci, "C112", _n(inputs.get("mud_monthly")))
    write_cell(ci, "D112", _n(inputs.get("mud_pct")))
    write_cell(ci, "C116", _n(inputs.get("insurance_monthly")))
    write_cell(ci, "C120", _n(inputs.get("bookkeeping_monthly")))

    # ── REVENUE INPUTS ────────────────────────────────────────────────
    ri = wb["Revenue Inputs"]

    write_cell(ri, "B2", _s(inputs.get("timing_method")))
    write_cell(ri, "B3", _n(inputs.get("bem_period")))
    write_cell(ri, "B4", _n(inputs.get("bem_pct")))
    write_cell(ri, "B5", _n(inputs.get("brokerage_fees")))
    write_cell(ri, "B6", _n(inputs.get("lot_closing_costs")))

    # $/FF by year (rows 13–23, year 0–10)
    price_per_ff = inputs.get("price_per_ff") or {}
    for yr in range(11):
        row = 13 + yr
        val = price_per_ff.get(yr) or price_per_ff.get(str(yr))
        if val is not None:
            write_cell(ri, f"B{row}", _n(val))

    # Home table (rows 27–42, FF 25–100)
    lot_sizes = inputs.get("lot_sizes", [])
    ht_rows = [27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42]
    for i, row in enumerate(ht_rows):
        ls = lot_sizes[i] if i < len(lot_sizes) else {}
        write_cell(ri, f"B{row}", _n(ls.get("build_time")))
        write_cell(ri, f"C{row}", _n(ls.get("home_price")))
        write_cell(ri, f"D{row}", _n(ls.get("av_pct")))
        write_cell(ri, f"F{row}", _n(ls.get("premium_per_ff")))
        write_cell(ri, f"G{row}", _n(ls.get("escalation")))
        write_cell(ri, f"H{row}", _n(ls.get("fence_per_ff")))
        write_cell(ri, f"I{row}", _n(ls.get("marketing_fee")))
        write_cell(ri, f"K{row}", _n(ls.get("lot_av_pct")))
        write_cell(ri, f"M{row}", _n(ls.get("lot_tax_rate")))

    # Residential pods (rows 46–51)
    write_cell(ri, "A46", _n(inputs.get("res_pod_acreage")))
    write_cell(ri, "B46", _n(inputs.get("res_pod_count")))
    res_pods = inputs.get("res_pods", [])
    res_pod_rows = [46, 47, 48, 49, 50, 51]
    for i, row in enumerate(res_pod_rows):
        rp = res_pods[i] if i < len(res_pods) else {}
        write_cell(ri, f"F{row}", _n(rp.get("price_per_acre")))
        write_cell(ri, f"G{row}", _n(rp.get("closing_costs_pct") or rp.get("closing_costs")))
        write_cell(ri, f"H{row}", _n(rp.get("implied_lots_per_acre")))
        write_cell(ri, f"I{row}", _n(rp.get("impact_fee_per_lot")))
        write_cell(ri, f"K{row}", _n(rp.get("sale_period")))

    # Commercial pods (rows 55–60)
    write_cell(ri, "A55", _n(inputs.get("comm_pod_acreage")))
    write_cell(ri, "B55", _n(inputs.get("comm_pod_count")))
    comm_pods = inputs.get("comm_pods", [])
    comm_pod_rows = [55, 56, 57, 58, 59, 60]
    for i, row in enumerate(comm_pod_rows):
        cp = comm_pods[i] if i < len(comm_pods) else {}
        write_cell(ri, f"F{row}", _n(cp.get("price_per_sf")))
        write_cell(ri, f"G{row}", _n(cp.get("closing_costs_pct") or cp.get("closing_costs")))
        write_cell(ri, f"I{row}", _n(cp.get("sale_period")))
        write_cell(ri, f"J{row}", _n(cp.get("av_per_acre")))
        write_cell(ri, f"K{row}", _n(cp.get("av_delay_months")))

    # MUD & WCID (rows 64–65)
    mud = inputs.get("mud_bond") or {}
    wcid = inputs.get("wcid_bond") or {}
    write_cell(ri, "B64", _n(mud.get("toggle")))
    write_cell(ri, "C64", _n(mud.get("debt_ratio")))
    write_cell(ri, "D64", _n(mud.get("first_bond_period")))
    write_cell(ri, "E64", _n(mud.get("bond_interval")))
    write_cell(ri, "F64", _n(mud.get("pct_to_dev")))
    write_cell(ri, "G64", _n(mud.get("receivables_fee")))
    write_cell(ri, "B65", _n(wcid.get("toggle")))
    write_cell(ri, "C65", _n(wcid.get("debt_ratio")))
    write_cell(ri, "D65", _n(wcid.get("first_bond_period")))
    write_cell(ri, "E65", _n(wcid.get("bond_interval")))
    write_cell(ri, "F65", _n(wcid.get("pct_to_dev")))
    write_cell(ri, "G65", _n(wcid.get("receivables_fee")))

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
