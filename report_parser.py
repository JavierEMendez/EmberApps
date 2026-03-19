"""
report_parser.py — Parses the Ember Dashboard Excel file and extracts
Consolidated Project Returns and Loan Capacities & Debt Schedules into
structured JSON.
"""

import io
from datetime import date, datetime
from typing import Any

import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _num(val: Any, precision: int | None = 2) -> float | int:
    """Convert a cell value to a number. None -> 0. Round if precision given."""
    if val is None:
        return 0
    try:
        n = float(val)
    except (TypeError, ValueError):
        return 0
    if precision is not None:
        n = round(n, precision)
    # Return int when the value is whole
    if precision is not None and n == int(n) and abs(n) < 1e15:
        return int(n)
    return n


def _str(val: Any) -> str:
    """Convert a cell value to a string. None -> ''."""
    if val is None:
        return ""
    return str(val).strip()


def _date_iso(val: Any) -> str:
    """Convert a cell value to an ISO-format date string."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    return s


def _row_yearly(ws, row: int, col_start: int = 7, col_end: int = 28,
                precision: int | None = 2) -> list:
    """Read yearly values from columns G(7) through AB(28)."""
    return [_num(ws.cell(row=row, column=c).value, precision)
            for c in range(col_start, col_end + 1)]


# ---------------------------------------------------------------------------
# Consolidated Project Returns
# ---------------------------------------------------------------------------

_PROJECT_STARTS = [4, 15, 26, 37, 48, 59, 70, 81, 92]

_METRIC_OFFSETS = [
    # (row offset from block start, label, has_status flag, precision)
    (2, "Preferred Return", True, 2),
    (3, "Return of Capital", True, 2),
    (4, "Excess Cash Flow", True, 2),
    (5, "Total LP Distributions", False, 2),
    (6, "Total LP Contributions", False, 2),
    (7, "Total LP Profit", False, 2),
    (8, "LP IRR", False, None),        # full precision for IRR
    (9, "LP Equity Multiple", False, None),  # full precision
    (10, "Promote", True, 2),
]


def _parse_returns(ws) -> dict:
    """Parse the 'Consolidated Project Returns' tab."""
    title = _str(ws.cell(row=3, column=3).value)  # C3
    report_date = _date_iso(ws.cell(row=3, column=26).value)  # Z3

    # Year headers from first project block row, columns G-AB
    years = []
    for c in range(7, 29):
        v = ws.cell(row=3, column=c).value
        if v is None:
            # Try row 4 (the first project header row) as fallback
            v = ws.cell(row=4, column=c).value
        try:
            yr = int(v)
            years.append(yr)
        except (TypeError, ValueError):
            # Fill gaps — likely a merged cell; infer from neighbors
            if years:
                years.append(years[-1] + 1)
            else:
                years.append(0)

    # --- Projects ---
    projects = []
    for start_row in _PROJECT_STARTS:
        name = _str(ws.cell(row=start_row, column=3).value)  # C column
        if not name:
            continue

        metrics = []
        for offset, label, has_status, prec in _METRIC_OFFSETS:
            r = start_row + offset
            metric: dict[str, Any] = {"label": _str(ws.cell(row=r, column=3).value) or label}
            if has_status:
                metric["status"] = _str(ws.cell(row=r, column=4).value)
            else:
                metric["status"] = _num(ws.cell(row=r, column=4).value, 0)
            metric["total"] = _num(ws.cell(row=r, column=5).value, prec)
            metric["yearly"] = _row_yearly(ws, r, precision=prec)

            metrics.append(metric)

        projects.append({"name": name, "metrics": metrics})

    # --- Summary section (rows 105-111) ---
    summary_labels = [
        (105, "MPC Contributions"),
        (106, "MPC Distributions"),
        (107, "MPC Net Cashflow"),
        (108, "Vertical Contributions"),
        (109, "Vertical Distributions"),
        (110, "Vertical Net Cashflow"),
        (111, "Total Assets Net Cashflow"),
    ]
    summary = []
    for r, default_label in summary_labels:
        label = _str(ws.cell(row=r, column=3).value) or default_label
        total = _num(ws.cell(row=r, column=5).value, 2)
        yearly = _row_yearly(ws, r, precision=2)
        summary.append({"label": label, "total": total, "yearly": yearly})

    return {
        "title": title or "Consolidated Ember Project Returns",
        "date": report_date,
        "years": years,
        "projects": projects,
        "summary": summary,
    }


# ---------------------------------------------------------------------------
# Loan Capacities & Debt Schedules
# ---------------------------------------------------------------------------

_LOAN_HEADERS = [
    "Community", "Lender", "Collateral", "Recourse",
    "Loan Origination", "Loan Term Date", "Months Remaining",
    "Rem. Interest Reserve", "Monthly Interest Burn",
    "Remaining Mos. of IR", "IR Health", "Index + Spread",
    "Today's Rate", "Extensions Remaining", "Extension Cost",
    "Loan Amount", "Drawn", "Balance", "Utilization",
    "Remaining", "Forecasted Thru Term", "Capacity Health",
]

# Column mapping: B=2, D=4, E=5, F=6, G=7, H=8, ... X=24
_LOAN_COLS = [2, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24]

# Fields that should be treated as dates
_DATE_FIELDS = {"Loan Origination", "Loan Term Date"}
# Fields that are percentages / rates — keep full precision
_RATE_FIELDS = {"Index + Spread", "Today's Rate", "Utilization"}
# Fields that are strings
_STR_FIELDS = {"Community", "Lender", "Collateral", "Recourse", "IR Health",
               "Capacity Health"}


def _read_loan_cell(header: str, val: Any) -> Any:
    if header in _DATE_FIELDS:
        return _date_iso(val)
    if header in _STR_FIELDS:
        return _str(val)
    if header in _RATE_FIELDS:
        return _num(val, precision=None)
    # Numeric by default
    return _num(val, 2)


def _read_loan_row(ws, row: int) -> dict:
    """Read a single loan data row into a dict keyed by header name."""
    result = {}
    for header, col in zip(_LOAN_HEADERS, _LOAN_COLS):
        result[header] = _read_loan_cell(header, ws.cell(row=row, column=col).value)
    return result


def _read_totals_row(ws, row: int) -> dict:
    """Read a totals row — only non-empty/non-zero fields."""
    totals = {}
    for header, col in zip(_LOAN_HEADERS, _LOAN_COLS):
        val = ws.cell(row=row, column=col).value
        if val is not None and val != "" and val != 0:
            totals[header] = _read_loan_cell(header, val)
    return totals


def _parse_loans(ws) -> dict:
    """Parse the 'Loan Capacities & DS' tab."""

    # Determine the report date from context (use today as fallback)
    report_date = date.today().isoformat()

    # --- MPC Loans (rows 7-10, totals row 11) ---
    mpc_rows = []
    for r in range(7, 11):
        val = ws.cell(row=r, column=2).value  # B column = Community
        if val is None or _str(val) == "":
            continue
        mpc_rows.append(_read_loan_row(ws, r))
    mpc_totals = _read_totals_row(ws, 11)

    # --- Vertical Loans (row 15 data, row 16 totals, row 17 footnote) ---
    vert_rows = []
    r = 15
    while True:
        val = ws.cell(row=r, column=2).value
        label = _str(val)
        if label == "" or label.lower() == "totals":
            break
        vert_rows.append(_read_loan_row(ws, r))
        r += 1
    vert_totals = _read_totals_row(ws, 16)
    footnote = _str(ws.cell(row=17, column=2).value)

    # --- Debt Schedules (starting at row 56) ---
    debt_schedules = _parse_debt_schedules(ws)

    return {
        "date": report_date,
        "mpc_loans": {
            "headers": list(_LOAN_HEADERS),
            "rows": mpc_rows,
            "totals": mpc_totals,
        },
        "vertical_loans": {
            "headers": list(_LOAN_HEADERS),
            "rows": vert_rows,
            "totals": vert_totals,
            "footnote": footnote,
        },
        "debt_schedules": debt_schedules,
    }


def _parse_debt_schedules(ws) -> list[dict]:
    """Parse debt schedule blocks starting at row 56."""
    schedules = []

    # First (and possibly only) debt schedule block
    project_name = _str(ws.cell(row=57, column=2).value)  # B57
    if not project_name:
        project_name = _str(ws.cell(row=57, column=3).value)  # C57 fallback
    # Clean project name — might contain extra text like "Associated Revenues"
    if project_name:
        project_name = project_name.split("\n")[0].strip()

    # Monthly date headers in L57-W57 (cols 12-23)
    months = []
    for c in range(12, 24):
        val = ws.cell(row=57, column=c).value
        months.append(_date_iso(val))

    # Payment rows 59-63
    payments = []
    for r in range(59, 64):
        dt = ws.cell(row=r, column=2).value   # B = date
        lender = ws.cell(row=r, column=3).value  # C = lender
        amount = ws.cell(row=r, column=4).value  # D = amount
        covered = ws.cell(row=r, column=5).value  # E = covered status
        if dt is None and lender is None and amount is None:
            continue
        payments.append({
            "date": _date_iso(dt),
            "lender": _str(lender),
            "amount": _num(amount, 2),
            "covered": _str(covered),
        })

    # Payment total — row 64, column D
    payment_total = _num(ws.cell(row=64, column=4).value, 2)

    # Revenue rows 60-67 (H=type, J=pct, K=total, L-W=monthly)
    # These overlap with payment rows in different columns
    revenues = []
    for r in range(60, 68):
        rev_type = ws.cell(row=r, column=8).value  # H
        if rev_type is None or _str(rev_type) == "":
            continue
        pct = _num(ws.cell(row=r, column=10).value, precision=None)  # J — keep full
        total = _num(ws.cell(row=r, column=11).value, 2)  # K
        monthly = [_num(ws.cell(row=r, column=c).value, 2) for c in range(12, 24)]
        revenues.append({
            "type": _str(rev_type),
            "pct": pct,
            "total": total,
            "monthly": monthly,
        })

    # Total Revenues — row 68
    total_rev_pct = _num(ws.cell(row=68, column=10).value, precision=None)
    total_rev_total = _num(ws.cell(row=68, column=11).value, 2)
    total_rev_monthly = [_num(ws.cell(row=68, column=c).value, 2) for c in range(12, 24)]
    total_revenues = {
        "pct": total_rev_pct,
        "total": total_rev_total,
        "monthly": total_rev_monthly,
    }

    # Cumulative Revenues — row 69, cols L-W
    cumulative_revenues = [_num(ws.cell(row=69, column=c).value, 2) for c in range(12, 24)]

    # Cumulative Payments — row 70, cols L-W
    cumulative_payments = [_num(ws.cell(row=70, column=c).value, 2) for c in range(12, 24)]

    schedules.append({
        "project": project_name,
        "months": months,
        "payments": payments,
        "payment_total": payment_total,
        "revenues": revenues,
        "total_revenues": total_revenues,
        "cumulative_revenues": cumulative_revenues,
        "cumulative_payments": cumulative_payments,
    })

    return schedules


# ---------------------------------------------------------------------------
# Operations (Ember Operating Revenues)
# ---------------------------------------------------------------------------

_OPS_CATEGORIES = [
    "Development Fees",
    "Project Personnel",
    "Bookkeeping",
    "Receivables & Bond Fees",
    "Ember Brokerage Fees",
]

def _parse_operations(ws) -> dict:
    """Parse the 'Operations' tab."""
    from datetime import date as _date

    # --- Determine column extent from row 52 (Model Dates) ---
    dates = []       # list of ISO date strings
    date_cols = []   # corresponding column indices
    for c in range(5, 300):
        v = ws.cell(row=52, column=c).value
        if v is None:
            break
        d = v.date() if hasattr(v, 'date') else v
        dates.append(d.isoformat() if hasattr(d, 'isoformat') else str(d))
        date_cols.append(c)

    if not dates:
        return {}

    # --- KPIs from D83:E85 ---
    kpis = []
    for r in range(83, 86):
        label = _str(ws.cell(row=r, column=4).value)
        value = _num(ws.cell(row=r, column=5).value, 2)
        kpis.append({"label": label, "value": value})

    # --- Expected Next 12 Months (sum of row 73 for next 12 months from today) ---
    today = _date.today()
    today_col = None
    for i, c in enumerate(date_cols):
        v = ws.cell(row=52, column=c).value
        d = v.date() if hasattr(v, 'date') else v
        if hasattr(d, 'year') and d.year == today.year and d.month == today.month:
            today_col = c
            break

    next_12_sum = 0
    if today_col:
        for c in range(today_col, min(today_col + 12, date_cols[-1] + 1)):
            val = ws.cell(row=73, column=c).value
            next_12_sum += _num(val, 2)
    kpis.append({"label": "Expected Next 12 Months", "value": round(next_12_sum, 2)})

    # --- Monthly data: rows 53-72 (per-project per-category) + row 73 totals ---
    # Structure: blocks of 5 rows per project, project name in col C of first row
    monthly_rows = []
    r = 53
    while r <= 72:
        project_name = _str(ws.cell(row=r, column=3).value)
        if not project_name:
            r += 1
            continue
        # 5 category rows per project
        for offset in range(5):
            cat = _str(ws.cell(row=r + offset, column=4).value)
            values = [_num(ws.cell(row=r + offset, column=c).value, 2)
                      for c in date_cols]
            monthly_rows.append({
                "project": project_name,
                "category": cat,
                "values": values,
            })
        r += 5

    # Row 73 totals
    monthly_totals = [_num(ws.cell(row=73, column=c).value, 2) for c in date_cols]

    # --- Yearly rollup: find next 5 years of data ---
    # Row 50 has years, rows 76-80 have category data, row 81 totals
    # Aggregate months by year
    year_map = {}  # year -> list of column indices
    for c in date_cols:
        yr = ws.cell(row=50, column=c).value
        if yr is None:
            continue
        yr = int(yr)
        year_map.setdefault(yr, []).append(c)

    # Next 5 calendar years starting from current year
    current_year = today.year
    yearly_years = [y for y in sorted(year_map.keys()) if y >= current_year][:5]
    yearly_rows = []
    for cat_row, cat_name in zip(range(76, 81), _OPS_CATEGORIES):
        values = []
        for yr in yearly_years:
            total = sum(_num(ws.cell(row=cat_row, column=c).value, 2)
                        for c in year_map[yr])
            values.append(round(total, 2))
        yearly_rows.append({"label": cat_name, "values": values})

    yearly_totals = []
    for yr in yearly_years:
        total = sum(_num(ws.cell(row=81, column=c).value, 2) for c in year_map[yr])
        yearly_totals.append(round(total, 2))

    # --- Quarterly rollup: next 12 quarters from today ---
    # Row 75 has quarter labels like "Q1 2026", rows 76-80 categories, row 81 totals
    # Group columns by quarter label
    quarter_map = {}  # quarter_label -> list of columns
    quarter_order = []
    for c in date_cols:
        qlabel = _str(ws.cell(row=75, column=c).value)
        if not qlabel:
            continue
        if qlabel not in quarter_map:
            quarter_map[qlabel] = []
            quarter_order.append(qlabel)
        quarter_map[qlabel].append(c)

    # Find current quarter
    q_num = (today.month - 1) // 3 + 1
    current_q = f"Q{q_num} {today.year}"
    try:
        start_idx = quarter_order.index(current_q)
    except ValueError:
        start_idx = 0
    next_12_quarters = quarter_order[start_idx:start_idx + 12]

    quarterly_rows = []
    for cat_row, cat_name in zip(range(76, 81), _OPS_CATEGORIES):
        values = []
        for qlabel in next_12_quarters:
            total = sum(_num(ws.cell(row=cat_row, column=c).value, 2)
                        for c in quarter_map[qlabel])
            values.append(round(total, 2))
        quarterly_rows.append({"label": cat_name, "values": values})

    quarterly_totals = []
    for qlabel in next_12_quarters:
        total = sum(_num(ws.cell(row=81, column=c).value, 2)
                    for c in quarter_map[qlabel])
        quarterly_totals.append(round(total, 2))

    # --- Next 12 months data (rows 76-80, 81) ---
    next_12_dates = []
    next_12_month_rows = []
    if today_col:
        n12_cols = [c for c in range(today_col, min(today_col + 12, date_cols[-1] + 1))]
        next_12_dates = [dates[date_cols.index(c)] for c in n12_cols]
        for cat_row, cat_name in zip(range(76, 81), _OPS_CATEGORIES):
            values = [_num(ws.cell(row=cat_row, column=c).value, 2) for c in n12_cols]
            next_12_month_rows.append({"label": cat_name, "values": values})
        n12_totals = [_num(ws.cell(row=81, column=c).value, 2) for c in n12_cols]
    else:
        n12_totals = []

    return {
        "kpis": kpis,
        "yearly_rollup": {
            "years": yearly_years,
            "rows": yearly_rows,
            "totals": yearly_totals,
        },
        "monthly": {
            "dates": dates,
            "rows": monthly_rows,
            "totals": monthly_totals,
        },
        "next_12_months": {
            "dates": next_12_dates,
            "rows": next_12_month_rows,
            "totals": n12_totals,
        },
        "quarterly_rollup": {
            "quarters": next_12_quarters,
            "rows": quarterly_rows,
            "totals": quarterly_totals,
        },
    }


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_dashboard(file_bytes: bytes) -> dict:
    """
    Parse an Ember Dashboard Excel file and return structured JSON with
    Consolidated Project Returns and Loan Capacities & Debt Schedules.

    Parameters
    ----------
    file_bytes : bytes
        Raw bytes of the .xlsx file.

    Returns
    -------
    dict
        Nested dictionary with "returns" and "loans" top-level keys.
    """
    wb = openpyxl.load_workbook(
        filename=io.BytesIO(file_bytes),
        data_only=True,
        read_only=False,
    )

    # --- Returns tab ---
    returns_ws = None
    for name in wb.sheetnames:
        if "consolidated" in name.lower() and "return" in name.lower():
            returns_ws = wb[name]
            break
    if returns_ws is None:
        # Fallback: try exact name
        returns_ws = wb.get("Consolidated Project Returns")

    returns_data = _parse_returns(returns_ws) if returns_ws else {}

    # --- Loans tab ---
    loans_ws = None
    for name in wb.sheetnames:
        if "loan" in name.lower() and ("capacit" in name.lower() or "ds" in name.lower()):
            loans_ws = wb[name]
            break
    if loans_ws is None:
        loans_ws = wb.get("Loan Capacities & DS")

    loans_data = _parse_loans(loans_ws) if loans_ws else {}

    # --- Operations tab ---
    ops_ws = None
    for name in wb.sheetnames:
        if "operation" in name.lower():
            ops_ws = wb[name]
            break
    if ops_ws is None:
        ops_ws = wb.get("Operations")

    ops_data = _parse_operations(ops_ws) if ops_ws else {}

    # Use the returns date for loans if available
    if returns_data.get("date") and loans_data:
        loans_data["date"] = returns_data["date"]

    wb.close()

    return {
        "returns": returns_data,
        "loans": loans_data,
        "operations": ops_data,
    }
