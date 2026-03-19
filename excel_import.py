"""
excel_import.py — reads an Ember underwriting Excel workbook and returns
an inputs dict that can be saved directly to a project.

Supports both the original template layout (exported by excel_export.py)
and the newer acquisitions model which has an extra header row in Cost Inputs,
shifting all data rows down by 1.
"""
import io
from typing import Any

import openpyxl


# ── helpers ──────────────────────────────────────────────────────────────

def _n(val: Any, default=0) -> float:
    """Safe numeric read — None/empty → default."""
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _s(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _int(val: Any, default=0) -> int:
    return int(_n(val, default))


def _is_numeric(val: Any) -> bool:
    """Check if a value is numeric (not a string header)."""
    if val is None:
        return True  # empty cell is fine
    try:
        float(val)
        return True
    except (TypeError, ValueError):
        return False


# ── main entry point ─────────────────────────────────────────────────────

def import_excel(file_bytes: bytes) -> dict:
    """
    Parse an Ember underwriting Excel file and return an inputs dict
    compatible with the project data model (same shape as default_inputs).
    """
    wb = openpyxl.load_workbook(
        filename=io.BytesIO(file_bytes),
        data_only=True,
        read_only=False,
    )

    # Find sheets by name (case-insensitive partial match)
    ti = _find_sheet(wb, "tract")
    ci = _find_sheet(wb, "cost")
    ri = _find_sheet(wb, "revenue")

    if not ti or not ci or not ri:
        missing = []
        if not ti:
            missing.append("Tract Inputs")
        if not ci:
            missing.append("Cost Inputs")
        if not ri:
            missing.append("Revenue Inputs")
        raise ValueError(f"Missing required sheets: {', '.join(missing)}. Found: {wb.sheetnames}")

    inputs = {}

    # ── TRACT INPUTS ─────────────────────────────────────────────────
    inputs["project_name"] = _s(ti["B5"].value) or "Imported Project"
    inputs["address"] = _s(ti["B6"].value)
    inputs["gross_acreage"] = _n(ti["B7"].value)
    inputs["land_escalator"] = _n(ti["B8"].value)
    inputs["purchase_price_per_acre"] = _n(ti["B9"].value)
    inputs["closing_costs_pct"] = _n(ti["B11"].value)
    inputs["closing_date"] = _s(ti["B14"].value)

    # Plants (rows 19-26)
    plants = []
    for row in range(19, 27):
        plants.append({
            "type": _s(ti.cell(row=row, column=2).value) or "None",
            "notes": _s(ti.cell(row=row, column=4).value),
        })
    inputs["plants"] = plants

    # Detention
    inputs["det_storage_rate"] = _n(ti["B31"].value, 1.1)
    inputs["det_depth"] = _n(ti["B33"].value, 9)
    inputs["det_num_projects"] = _int(ti["B34"].value, 6)

    # Amenities (rows 42-47)
    amenities = []
    for row in range(42, 48):
        amenities.append({
            "type": _s(ti.cell(row=row, column=2).value) or "None",
            "acres": _n(ti.cell(row=row, column=3).value),
        })
    inputs["amenities"] = amenities

    # Parks / drill sites
    inputs["parks_pct"] = _n(ti["B51"].value, 0.03)
    inputs["drill_site_acres"] = _n(ti["C52"].value)

    # Other net-outs (rows 56-61)
    other_netouts = []
    for row in range(56, 62):
        other_netouts.append({
            "description": _s(ti.cell(row=row, column=1).value),
            "acres": _n(ti.cell(row=row, column=2).value),
        })
    inputs["other_netouts"] = other_netouts

    # Roads (rows 65-70)
    roads = []
    for row in range(65, 71):
        roads.append({
            "type": _s(ti.cell(row=row, column=2).value),
            "linear_feet": _n(ti.cell(row=row, column=3).value),
            "width": _n(ti.cell(row=row, column=4).value),
            "road_setback": _n(ti.cell(row=row, column=5).value),
            "landscaping_setback": _n(ti.cell(row=row, column=6).value),
        })
    inputs["roads"] = roads

    # Pod acres
    inputs["commercial_pod_acres"] = _n(ti["B76"].value)
    inputs["residential_pod_acres"] = _n(ti["B77"].value)

    # ── COST INPUTS ──────────────────────────────────────────────────
    # Auto-detect layout: in the original template, D25 is a number (plant
    # base cost). In the newer model, D25 is a header string like "Base Cost ($)"
    # and data starts one row later. Same +1 shift applies throughout.
    _v25 = ci.cell(row=25, column=4).value  # D25
    off = 1 if not _is_numeric(_v25) else 0  # offset for new model

    inputs["default_other_pct"] = _n(ci["B5"].value, 0.17)
    inputs["sectional_other_pct"] = _n(ci["B6"].value, 0.17)
    inputs["landscaping_other_pct"] = _n(ci["B7"].value, 0.12)
    inputs["contingency"] = _n(ci["B8"].value, 0.05)
    inputs["site_work_pct"] = _n(ci["B9"].value, 0.01)
    inputs["fenced_pct"] = _n(ci["B10"].value, 0.25)
    inputs["cost_per_mailbox"] = _n(ci["B11"].value, 200)
    inputs["cost_per_streetlight"] = _n(ci["B12"].value, 1700)
    inputs["default_start_month"] = _int(ci["B13"].value, 1)

    # Takedowns (cols B/C/D — period at row 17, pct at row 18+off)
    # In the new model, row 18 = acreage, row 19 = pct (off=1)
    # In the old template, row 18 = pct directly (off=0)
    td_pct_row = 18 + off
    takedowns = []
    for col_letter in ["B", "C", "D"]:
        takedowns.append({
            "period": _int(ci[f"{col_letter}17"].value),
            "pct": _n(ci.cell(row=td_pct_row, column=ord(col_letter) - 64).value),
        })
    inputs["takedowns"] = takedowns

    # Plant costs (rows 25-32 + off)
    plant_costs = []
    for row in range(25 + off, 33 + off):
        plant_costs.append({
            "base_cost": _n(ci.cell(row=row, column=4).value),       # D
            "other_pct": _n(ci.cell(row=row, column=5).value, 0.17), # E
            "start_month": _int(ci.cell(row=row, column=7).value, 1),# G
            "ph2_base_cost": _n(ci.cell(row=row, column=10).value),  # J
            "ph2_other_pct": _n(ci.cell(row=row, column=11).value, 0.17), # K
            "ph2_start_month": _int(ci.cell(row=row, column=13).value, 37), # M
        })
    inputs["plant_costs"] = plant_costs

    # Amenity costs (rows 36-41 + off)
    amenity_costs = []
    for row in range(36 + off, 42 + off):
        amenity_costs.append({
            "base_cost": _n(ci.cell(row=row, column=4).value),       # D
            "other_pct": _n(ci.cell(row=row, column=5).value, 0.17), # E
            "start_month": _int(ci.cell(row=row, column=7).value, 1),# G
        })
    inputs["amenity_costs"] = amenity_costs

    # Detention costs (rows 45-50 + off)
    det_costs = []
    for row in range(45 + off, 51 + off):
        det_costs.append({
            "other_pct": _n(ci.cell(row=row, column=4).value, 0.17), # D
            "start_month": _int(ci.cell(row=row, column=6).value, 1),# F
            "landscaping_per_foot": _n(ci.cell(row=row, column=9).value, 2), # I
        })
    inputs["det_costs"] = det_costs

    # Other costs (rows 54-59 + off)
    other_costs = []
    for row in range(54 + off, 60 + off):
        other_costs.append({
            "base_cost": _n(ci.cell(row=row, column=4).value),       # D
            "other_pct": _n(ci.cell(row=row, column=5).value, 0.17), # E
            "start_month": _int(ci.cell(row=row, column=7).value, 1),# G
            "duration": _int(ci.cell(row=row, column=8).value, 1),   # H
        })
    inputs["other_costs"] = other_costs

    # Road costs (rows 63-68 + off)
    road_costs = []
    for row in range(63 + off, 69 + off):
        road_costs.append({
            "wsd_per_lf": _n(ci.cell(row=row, column=4).value),       # D
            "paving_per_lf": _n(ci.cell(row=row, column=5).value),    # E
            "other_pct": _n(ci.cell(row=row, column=7).value, 0.17),  # G
            "start_month": _int(ci.cell(row=row, column=9).value, 1), # I
            "landscaping_per_sf": _n(ci.cell(row=row, column=12).value, 2), # L
            "light_spacing": _n(ci.cell(row=row, column=14).value),   # N
        })
    inputs["road_costs"] = road_costs

    # Lot sizes (rows 72-87 + off, 16 sizes from 25FF to 100FF)
    lot_sizes = []
    ff_values = [25, 30, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85, 90, 95, 100]
    for idx, row in enumerate(range(72 + off, 88 + off)):
        ls = {
            "front_footage": ff_values[idx],
            "on": _int(ci.cell(row=row, column=2).value),             # B
            "yield_per_ac": _n(ci.cell(row=row, column=3).value),     # C
            "pace": _n(ci.cell(row=row, column=4).value),             # D
            "home_price": _n(ci.cell(row=row, column=8).value),       # H
            "wsd_per_ff": _n(ci.cell(row=row, column=9).value),       # I
            "paving_per_ff": _n(ci.cell(row=row, column=10).value),   # J
            "dev_start_month": _int(ci.cell(row=row, column=12).value, 1), # L
            "landscaping_per_lot": _n(ci.cell(row=row, column=13).value),  # M
            "urd_per_lot": _n(ci.cell(row=row, column=14).value),     # N
            "lots_per_streetlight": _int(ci.cell(row=row, column=15).value, 4), # O
            "fence_cost_per_ff": _n(ci.cell(row=row, column=16).value),    # P
        }
        lot_sizes.append(ls)

    # Overhead (rows shifted by off)
    inputs["prof_svc_pct"] = _n(ci.cell(row=95 + off, column=2).value, 0.015)      # B95/96
    inputs["dmf_pct"] = _n(ci.cell(row=99 + off, column=2).value, 0.025)           # B99/100
    inputs["personnel_monthly"] = _n(ci.cell(row=103 + off, column=3).value, 50000) # C103/104
    inputs["marketing_personnel_monthly"] = _n(ci.cell(row=104 + off, column=3).value, 15000) # C104/105
    inputs["legal_monthly"] = _n(ci.cell(row=108 + off, column=3).value, 10000)     # C108/109
    inputs["mud_monthly"] = _n(ci.cell(row=112 + off, column=3).value, 35000)       # C112/113
    inputs["mud_pct"] = _n(ci.cell(row=112 + off, column=4).value, 0.2)            # D112/113
    inputs["insurance_monthly"] = _n(ci.cell(row=116 + off, column=3).value, 10000) # C116/117
    inputs["bookkeeping_monthly"] = _n(ci.cell(row=120 + off, column=3).value, 10000) # C120/121

    # ── REVENUE INPUTS ───────────────────────────────────────────────
    inputs["timing_method"] = _s(ri["B2"].value) or "50/25/25"
    inputs["bem_period"] = _int(ri["B3"].value, 9)
    inputs["bem_pct"] = _n(ri["B4"].value, 0.18)
    inputs["brokerage_fees"] = _n(ri["B5"].value, 0.03)
    inputs["lot_closing_costs"] = _n(ri["B6"].value, 0.015)

    # Take weights (new model has these at B7/B8/B9)
    take1 = _n(ri["B7"].value)
    take2 = _n(ri["B8"].value)
    take3 = _n(ri["B9"].value)

    # $/FF by year (rows 13-23)
    price_per_ff = {}
    for yr in range(11):
        val = _n(ri.cell(row=13 + yr, column=2).value)
        price_per_ff[str(yr)] = val
    inputs["price_per_ff"] = price_per_ff

    # Home build table — merge into lot_sizes (rows 27-42)
    for idx, row in enumerate(range(27, 43)):
        if idx < len(lot_sizes):
            lot_sizes[idx]["build_time"] = _int(ri.cell(row=row, column=2).value)   # B
            hp = _n(ri.cell(row=row, column=3).value)
            if hp:
                lot_sizes[idx]["home_price"] = hp  # C (override if present)
            lot_sizes[idx]["av_pct"] = _n(ri.cell(row=row, column=4).value)          # D
            lot_sizes[idx]["premium_per_ff"] = _n(ri.cell(row=row, column=6).value)  # F
            lot_sizes[idx]["escalation"] = _n(ri.cell(row=row, column=7).value)      # G
            lot_sizes[idx]["fence_per_ff"] = _n(ri.cell(row=row, column=8).value)    # H
            lot_sizes[idx]["marketing_fee"] = _n(ri.cell(row=row, column=9).value)   # I
            lot_sizes[idx]["lot_av_pct"] = _n(ri.cell(row=row, column=11).value)     # K
            lot_sizes[idx]["lot_tax_rate"] = _n(ri.cell(row=row, column=13).value)   # M

    inputs["lot_sizes"] = lot_sizes

    # Residential pods (rows 46-51)
    inputs["res_pod_acreage"] = _n(ri["A46"].value)
    inputs["res_pod_count"] = _int(ri["B46"].value, 1)
    res_pods = []
    for row in range(46, 52):
        res_pods.append({
            "price_per_acre": _n(ri.cell(row=row, column=6).value),          # F
            "closing_costs_pct": _n(ri.cell(row=row, column=7).value, 0.045),# G
            "implied_lots_per_acre": _n(ri.cell(row=row, column=8).value),   # H
            "impact_fee_per_lot": _n(ri.cell(row=row, column=9).value),      # I
            "sale_period": _int(ri.cell(row=row, column=11).value, 12),      # K
        })
    inputs["res_pods"] = res_pods

    # Commercial pods (rows 55-60)
    inputs["comm_pod_acreage"] = _n(ri["A55"].value)
    inputs["comm_pod_count"] = _int(ri["B55"].value, 6)
    comm_pods = []
    for row in range(55, 61):
        comm_pods.append({
            "price_per_sf": _n(ri.cell(row=row, column=6).value),            # F
            "closing_costs_pct": _n(ri.cell(row=row, column=7).value, 0.045),# G
            "sale_period": _int(ri.cell(row=row, column=9).value, 12),       # I
            "av_per_acre": _n(ri.cell(row=row, column=10).value),            # J
            "av_delay_months": _int(ri.cell(row=row, column=11).value, 18),  # K
        })
    inputs["comm_pods"] = comm_pods

    # MUD & WCID bonds (rows 64-65)
    inputs["mud_bond"] = {
        "toggle": _int(ri["B64"].value, 1),
        "debt_ratio": _n(ri["C64"].value, 0.12),
        "first_bond_period": _int(ri["D64"].value, 48),
        "bond_interval": _int(ri["E64"].value, 12),
        "pct_to_dev": _n(ri["F64"].value, 0.85),
        "receivables_fee": _n(ri["G64"].value, 0.025),
    }
    inputs["wcid_bond"] = {
        "toggle": _int(ri["B65"].value, 1),
        "debt_ratio": _n(ri["C65"].value, 0.042),
        "first_bond_period": _int(ri["D65"].value, 48),
        "bond_interval": _int(ri["E65"].value, 12),
        "pct_to_dev": _n(ri["F65"].value, 0.85),
        "receivables_fee": _n(ri["G65"].value, 0.025),
    }

    # Derived fields the app expects
    inputs["take1_pct"] = take1 if take1 else (takedowns[0]["pct"] if takedowns else 0.5)
    inputs["take2_pct"] = take2 if take2 else (takedowns[1]["pct"] if len(takedowns) > 1 else 0.25)
    inputs["take3_pct"] = take3 if take3 else (takedowns[2]["pct"] if len(takedowns) > 2 else 0.25)
    inputs["marketing_pct"] = 0.02  # not in Excel, use default

    wb.close()
    return inputs


def _find_sheet(wb, keyword: str):
    """Find a sheet whose name contains the keyword (case-insensitive)."""
    kw = keyword.lower()
    for name in wb.sheetnames:
        if kw in name.lower():
            return wb[name]
    return None
