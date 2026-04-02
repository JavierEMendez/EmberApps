"""
Ember Tract Underwriting Web App
Flask + PostgreSQL + Flask-Login — no Excel required
"""
import os, json, datetime, io, base64
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
from functools import wraps
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
import psycopg2
import psycopg2.extras
from werkzeug.security import generate_password_hash, check_password_hash
from calc import calculate
from report_parser import parse_dashboard

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "ember-dev-secret-change-in-production")

# Auto-initialize DB on first request
_db_initialized = False

@app.before_request
def auto_init():
    global _db_initialized
    if not _db_initialized:
        try:
            init_db()
            _db_initialized = True
        except Exception as e:
            print(f"DB init error: {e}")

# ─── DATABASE ────────────────────────────────────────────────────────────────
def get_db():
    conn = psycopg2.connect(os.environ["DATABASE_URL"], cursor_factory=psycopg2.extras.RealDictCursor)
    return conn

def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            is_admin BOOLEAN DEFAULT FALSE,
            page_access JSONB NOT NULL DEFAULT '{"mpc_underwriting":true,"returns":true,"loans":true,"operations":true}'::jsonb,
            created_at TIMESTAMP DEFAULT NOW()
        );
        -- Add columns if upgrading from older schema
        ALTER TABLE users ADD COLUMN IF NOT EXISTS page_access JSONB NOT NULL DEFAULT '{"mpc_underwriting":true,"returns":true,"loans":true,"operations":true}'::jsonb;
        ALTER TABLE users ADD COLUMN IF NOT EXISTS email TEXT;
        ALTER TABLE users ADD COLUMN IF NOT EXISTS report_opt_in BOOLEAN DEFAULT FALSE;
        ALTER TABLE users ADD COLUMN IF NOT EXISTS report_format TEXT DEFAULT 'pdf';
        CREATE TABLE IF NOT EXISTS report_sends (
            id SERIAL PRIMARY KEY,
            period TEXT UNIQUE NOT NULL,
            sent_at TIMESTAMP DEFAULT NOW()
        );
        ALTER TABLE projects ADD COLUMN IF NOT EXISTS scenarios JSONB DEFAULT '[]'::jsonb;
        ALTER TABLE projects ADD COLUMN IF NOT EXISTS status TEXT DEFAULT 'Active';
        ALTER TABLE projects ADD COLUMN IF NOT EXISTS change_log JSONB DEFAULT '[]'::jsonb;
        CREATE TABLE IF NOT EXISTS projects (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            address TEXT,
            created_by INTEGER REFERENCES users(id),
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW(),
            inputs JSONB NOT NULL DEFAULT '{}'::jsonb,
            outputs JSONB NOT NULL DEFAULT '{}'::jsonb,
            archived BOOLEAN DEFAULT FALSE
        );
        CREATE TABLE IF NOT EXISTS reports (
            id SERIAL PRIMARY KEY,
            report_type TEXT NOT NULL,
            data JSONB NOT NULL DEFAULT '{}'::jsonb,
            uploaded_by INTEGER REFERENCES users(id),
            uploaded_at TIMESTAMP DEFAULT NOW()
        );
    """)
    # Backfill portfolio access for existing users
    cur.execute("UPDATE users SET page_access = page_access || '{\"portfolio\": true}'::jsonb WHERE page_access->>'portfolio' IS NULL")
    # Create default admin if no users exist
    cur.execute("SELECT COUNT(*) as cnt FROM users")
    row = cur.fetchone()
    if row["cnt"] == 0:
        cur.execute(
            "INSERT INTO users (username, password_hash, is_admin) VALUES (%s, %s, TRUE)",
            ("admin", generate_password_hash("ember2024"))
        )
    conn.commit()
    cur.close()
    conn.close()

# ─── AUTH HELPERS ─────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            if request.is_json:
                return jsonify({"error": "Unauthorized"}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("is_admin"):
            return jsonify({"error": "Admin required"}), 403
        return f(*args, **kwargs)
    return decorated

# ─── AUTH ROUTES ─────────────────────────────────────────────────────────────
@app.route("/health")
def health():
    return "ok", 200

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE username = %s", (username,))
        user = cur.fetchone()
        cur.close(); conn.close()
        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["is_admin"] = user["is_admin"]
            session["page_access"] = user.get("page_access") or {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
            return redirect(url_for("home"))
        error = "Invalid username or password."
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ─── MAIN APP ─────────────────────────────────────────────────────────────────
@app.route("/home")
@login_required
def home():
    pa = session.get("page_access") or {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    # Admins always have full access
    if session.get("is_admin"):
        pa = {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    return render_template("home.html", username=session.get("username"), is_admin=session.get("is_admin"), page_access=pa)

@app.route("/")
@login_required
def index():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("mpc_underwriting", True):
        return redirect(url_for("home"))
    pa = session.get("page_access") or {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    if session.get("is_admin"):
        pa = {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    return render_template("app.html", username=session.get("username"), is_admin=session.get("is_admin"), page_access=pa)

# ─── PROJECT API ─────────────────────────────────────────────────────────────
@app.route("/api/projects", methods=["GET"])
@login_required
def list_projects():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT p.id, p.name, p.address, p.updated_at,
               u.username as created_by,
               p.outputs->>'total_revenue' as total_revenue,
               p.outputs->>'gross_margin_pct' as gross_margin_pct,
               p.outputs->>'total_lots' as total_lots,
               p.outputs->>'unlevered_irr' as unlevered_irr,
               p.outputs->>'project_length_years' as project_length_years,
               p.archived,
               COALESCE(p.status, 'Active') as status
        FROM projects p
        LEFT JOIN users u ON p.created_by = u.id
        WHERE p.archived = FALSE
        ORDER BY p.updated_at DESC
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/projects", methods=["POST"])
@login_required
def create_project():
    data = request.json or {}
    name = data.get("name", "New Project")
    inputs = default_inputs(name)
    try:
        outputs = calculate(inputs)
    except Exception:
        outputs = {}
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO projects (name, address, created_by, inputs, outputs) VALUES (%s, %s, %s, %s, %s) RETURNING id",
        (name, data.get("address", ""), session["user_id"], json.dumps(inputs), json.dumps(outputs))
    )
    pid = cur.fetchone()["id"]
    conn.commit(); cur.close(); conn.close()
    return jsonify({"id": pid, "name": name})

@app.route("/api/projects/<int:pid>", methods=["GET"])
@login_required
def get_project(pid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return jsonify({"error": "Not found"}), 404
    return jsonify(dict(row))

@app.route("/api/projects/<int:pid>", methods=["PUT"])
@login_required
def save_project(pid):
    data = request.json or {}
    inputs = data.get("inputs", {})
    try:
        outputs = calculate(inputs)
    except Exception as e:
        return jsonify({"error": f"Calculation error: {e}"}), 500
    conn = get_db()
    cur = conn.cursor()
    # Build change log entry
    cur.execute("SELECT inputs, change_log FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    old_log = list(row["change_log"] or []) if row else []
    changes = _compare_inputs(row["inputs"] or {} if row else {}, inputs)
    if changes:
        old_log.append({
            "ts": datetime.datetime.utcnow().isoformat() + "Z",
            "user": session.get("username", "unknown"),
            "changes": changes
        })
        old_log = old_log[-200:]
    cur.execute("""
        UPDATE projects
        SET inputs = %s, outputs = %s, name = %s, address = %s,
            change_log = %s, updated_at = NOW()
        WHERE id = %s
    """, (
        json.dumps(inputs),
        json.dumps(outputs),
        inputs.get("project_name", "Unnamed"),
        inputs.get("address", ""),
        json.dumps(old_log),
        pid
    ))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "outputs": outputs})

@app.route("/api/projects/<int:pid>", methods=["DELETE"])
@login_required
def delete_project(pid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE projects SET archived = TRUE WHERE id = %s", (pid,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/projects/<int:pid>/status", methods=["PATCH"])
@login_required
def set_project_status(pid):
    data = request.json or {}
    status = data.get("status", "Active")
    if status not in {"Active", "Under Contract", "Closed", "Dead"}:
        return jsonify({"error": "Invalid status"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE projects SET status = %s WHERE id = %s", (status, pid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/projects/<int:pid>/changelog", methods=["GET"])
@login_required
def get_changelog(pid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT change_log FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row: return jsonify({"error": "Not found"}), 404
    return jsonify(list(reversed(row["change_log"] or [])))

@app.route("/api/projects/<int:pid>/sensitivity", methods=["POST"])
@login_required
def sensitivity(pid):
    data = request.json or {}
    axis_x = data.get("axis_x", {})
    axis_y = data.get("axis_y", {})
    base_inputs = data.get("base_inputs", {})
    x_field, x_vals = axis_x.get("field"), axis_x.get("values", [])
    y_field, y_vals = axis_y.get("field"), axis_y.get("values", [])
    if not x_field or not y_field or not x_vals or not y_vals:
        return jsonify({"error": "Missing axis config"}), 400
    # Cap grid size for performance
    x_vals = x_vals[:7]
    y_vals = y_vals[:7]
    matrix = []
    for yv in y_vals:
        row_results = []
        for xv in x_vals:
            inp = _apply_sensitivity_override(base_inputs, x_field, xv)
            inp = _apply_sensitivity_override(inp, y_field, yv)
            try:
                out = calculate(inp)
                row_results.append({
                    "irr": out.get("unlevered_irr"),
                    "gm_pct": out.get("gross_margin_pct"),
                })
            except Exception:
                row_results.append({"irr": None, "gm_pct": None})
        matrix.append(row_results)
    return jsonify({"ok": True, "matrix": matrix, "x_values": x_vals, "y_values": y_vals})

# ─── SCENARIO API ─────────────────────────────────────────────────────────────
@app.route("/api/projects/<int:pid>/scenarios", methods=["GET"])
@login_required
def list_scenarios(pid):
    import uuid
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT inputs, outputs, scenarios FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    scenarios = list(row["scenarios"] or [])
    if not scenarios:
        sid = str(uuid.uuid4())[:8]
        scenarios = [{"id": sid, "name": "Base Case",
                      "inputs": row["inputs"] or {}, "outputs": row["outputs"] or {}}]
        cur.execute("UPDATE projects SET scenarios = %s WHERE id = %s", (json.dumps(scenarios), pid))
        conn.commit()
    cur.close(); conn.close()
    return jsonify(scenarios)

@app.route("/api/projects/<int:pid>/scenarios", methods=["POST"])
@login_required
def create_scenario(pid):
    import uuid
    data = request.json or {}
    name = data.get("name", "New Scenario").strip() or "New Scenario"
    clone_id = data.get("clone_from")
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT inputs, outputs, scenarios FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    scenarios = list(row["scenarios"] or [])
    sid = str(uuid.uuid4())[:8]
    if clone_id:
        src = next((s for s in scenarios if s["id"] == clone_id), None)
        inp = dict(src["inputs"]) if src else dict(row["inputs"] or {})
        out = dict(src["outputs"]) if src else dict(row["outputs"] or {})
    else:
        inp = dict(row["inputs"] or {})
        out = dict(row["outputs"] or {})
    new_scen = {"id": sid, "name": name, "inputs": inp, "outputs": out}
    scenarios.append(new_scen)
    cur.execute("UPDATE projects SET scenarios = %s WHERE id = %s", (json.dumps(scenarios), pid))
    conn.commit(); cur.close(); conn.close()
    return jsonify(new_scen)

@app.route("/api/projects/<int:pid>/scenarios/<sid>", methods=["PUT"])
@login_required
def save_scenario(pid, sid):
    data = request.json or {}
    inp = data.get("inputs", {})
    try:
        out = calculate(inp)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT scenarios FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    scenarios = list(row["scenarios"] or [])
    idx = next((i for i, s in enumerate(scenarios) if s["id"] == sid), None)
    if idx is None: cur.close(); conn.close(); return jsonify({"error": "Scenario not found"}), 404
    scenarios[idx]["inputs"] = inp
    scenarios[idx]["outputs"] = out
    cur.execute("UPDATE projects SET scenarios = %s WHERE id = %s", (json.dumps(scenarios), pid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "outputs": out})

@app.route("/api/projects/<int:pid>/scenarios/<sid>", methods=["DELETE"])
@login_required
def delete_scenario(pid, sid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT scenarios FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    scenarios = [s for s in (row["scenarios"] or []) if s["id"] != sid]
    cur.execute("UPDATE projects SET scenarios = %s WHERE id = %s", (json.dumps(scenarios), pid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/projects/<int:pid>/scenarios/<sid>/promote", methods=["POST"])
@login_required
def promote_scenario(pid, sid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT scenarios FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    scen = next((s for s in (row["scenarios"] or []) if s["id"] == sid), None)
    if not scen: cur.close(); conn.close(); return jsonify({"error": "Scenario not found"}), 404
    inp = scen["inputs"]; out = scen["outputs"]
    cur.execute("""UPDATE projects SET inputs=%s, outputs=%s, name=%s, address=%s, updated_at=NOW()
                   WHERE id=%s""",
                (json.dumps(inp), json.dumps(out),
                 inp.get("project_name", "Unnamed"), inp.get("address", ""), pid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "outputs": out})

@app.route("/api/projects/<int:pid>/scenarios/<sid>/name", methods=["PATCH"])
@login_required
def rename_scenario(pid, sid):
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name: return jsonify({"error": "Name required"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT scenarios FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    scenarios = list(row["scenarios"] or [])
    idx = next((i for i, s in enumerate(scenarios) if s["id"] == sid), None)
    if idx is None: cur.close(); conn.close(); return jsonify({"error": "Scenario not found"}), 404
    scenarios[idx]["name"] = name
    cur.execute("UPDATE projects SET scenarios = %s WHERE id = %s", (json.dumps(scenarios), pid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/projects/<int:pid>/calculate", methods=["POST"])
@login_required
def recalculate(pid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT inputs FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        return jsonify({"error": "Not found"}), 404
    inputs = row["inputs"]
    try:
        outputs = calculate(inputs)
    except Exception as e:
        cur.close(); conn.close()
        return jsonify({"error": str(e)}), 500
    cur.execute("UPDATE projects SET outputs = %s, updated_at = NOW() WHERE id = %s",
                (json.dumps(outputs), pid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "outputs": outputs})

# ─── ADMIN API ────────────────────────────────────────────────────────────────
@app.route("/api/admin/users", methods=["GET"])
@login_required
@admin_required
def list_users():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, username, email, is_admin, page_access, created_at, report_opt_in, report_format FROM users ORDER BY id")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/admin/users", methods=["POST"])
@login_required
@admin_required
def create_user():
    data = request.json or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    email = data.get("email", "").strip() or None
    is_admin = data.get("is_admin", False)
    page_access = data.get("page_access", {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True})
    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO users (username, password_hash, email, is_admin, page_access) VALUES (%s, %s, %s, %s, %s) RETURNING id",
            (username, generate_password_hash(password), email, is_admin, json.dumps(page_access))
        )
        uid = cur.fetchone()["id"]
        conn.commit()
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        cur.close(); conn.close()
        return jsonify({"error": "Username already exists"}), 409
    cur.close(); conn.close()
    return jsonify({"id": uid, "username": username})

@app.route("/api/admin/users/<int:uid>", methods=["DELETE"])
@login_required
@admin_required
def delete_user(uid):
    if uid == session["user_id"]:
        return jsonify({"error": "Cannot delete yourself"}), 400
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM users WHERE id = %s", (uid,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/admin/users/<int:uid>/password", methods=["PUT"])
@login_required
@admin_required
def reset_password(uid):
    data = request.json or {}
    password = data.get("password", "")
    if not password:
        return jsonify({"error": "Password required"}), 400
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE users SET password_hash = %s WHERE id = %s",
                (generate_password_hash(password), uid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/account", methods=["GET"])
@login_required
def get_account():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT username, email, report_opt_in, report_format FROM users WHERE id = %s", (session["user_id"],))
    row = cur.fetchone()
    cur.close(); conn.close()
    return jsonify({
        "username": row["username"],
        "email": row["email"] or "",
        "report_opt_in": bool(row["report_opt_in"]),
        "report_format": row["report_format"] or "pdf"
    })

@app.route("/api/admin/users/<int:uid>/email", methods=["PUT"])
@login_required
@admin_required
def set_user_email(uid):
    data = request.json or {}
    email = data.get("email", "").strip() or None
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE users SET email = %s WHERE id = %s", (email, uid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/account/password", methods=["PUT"])
@login_required
def change_own_password():
    data = request.json or {}
    current_pw = data.get("current_password", "")
    new_pw = data.get("new_password", "")
    if not current_pw or not new_pw:
        return jsonify({"error": "All fields are required"}), 400
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT password_hash FROM users WHERE id = %s", (session["user_id"],))
    user = cur.fetchone()
    if not user or not check_password_hash(user["password_hash"], current_pw):
        cur.close(); conn.close()
        return jsonify({"error": "Current password is incorrect"}), 400
    cur.execute("UPDATE users SET password_hash = %s WHERE id = %s",
                (generate_password_hash(new_pw), session["user_id"]))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/admin/users/<int:uid>/access", methods=["PUT"])
@login_required
@admin_required
def update_page_access(uid):
    data = request.json or {}
    page_access = data.get("page_access", {})
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE users SET page_access = %s WHERE id = %s",
                (json.dumps(page_access), uid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/account/report-settings", methods=["PUT"])
@login_required
def update_report_settings():
    data = request.json or {}
    opt_in = bool(data.get("report_opt_in", False))
    fmt = data.get("report_format", "pdf")
    if fmt not in ("pdf", "excel"):
        fmt = "pdf"
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE users SET report_opt_in = %s, report_format = %s WHERE id = %s",
                (opt_in, fmt, session["user_id"]))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/admin/send-reports-now", methods=["POST"])
@login_required
@admin_required
def send_reports_now():
    import threading
    def _run():
        try:
            count = _send_monthly_emails(force=True)
            print(f"[Reports] Sent to {count} recipient(s)", flush=True)
        except Exception as e:
            print(f"[Reports] Send failed: {e}", flush=True)
    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return jsonify({"ok": True, "queued": True})

# ─── DEFAULT INPUTS TEMPLATE ──────────────────────────────────────────────────
def default_inputs(name="New Project"):
    # Lot size defaults match Excel Cost Inputs rows 72-87 exactly
    # Columns: FF, on, yield/ac, pace lots/mo, home_price, wsd/ff, paving/ff, landscaping/lot, urd/lot, lots_per_streetlight, fence/ff
    lot_size_defaults = [
        {"front_footage":25,  "on":0, "yield_per_ac":8.25, "pace":5,    "home_price":200000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":30,  "on":0, "yield_per_ac":5.54, "pace":5,    "home_price":360000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":35,  "on":0, "yield_per_ac":8.25, "pace":6,    "home_price":275000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":40,  "on":1, "yield_per_ac":5.5,  "pace":7,    "home_price":330168,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":45,  "on":1, "yield_per_ac":5.0,  "pace":6,    "home_price":380000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":50,  "on":1, "yield_per_ac":4.5,  "pace":5,    "home_price":430000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":55,  "on":0, "yield_per_ac":4.0,  "pace":5,    "home_price":500000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":60,  "on":1, "yield_per_ac":3.5,  "pace":2,    "home_price":580000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":65,  "on":0, "yield_per_ac":3.0,  "pace":2,    "home_price":615000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":3, "fence_cost_per_ff":94},
        {"front_footage":70,  "on":0, "yield_per_ac":2.5,  "pace":1,    "home_price":675000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":3, "fence_cost_per_ff":94},
        {"front_footage":75,  "on":0, "yield_per_ac":2.0,  "pace":1,    "home_price":720000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":3, "fence_cost_per_ff":94},
        {"front_footage":80,  "on":1, "yield_per_ac":1.5,  "pace":0.75, "home_price":750000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":3, "fence_cost_per_ff":94},
        {"front_footage":85,  "on":0, "yield_per_ac":5.5,  "pace":0.75, "home_price":325000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":2, "fence_cost_per_ff":94},
        {"front_footage":90,  "on":0, "yield_per_ac":5.5,  "pace":0.75, "home_price":360000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":2, "fence_cost_per_ff":94},
        {"front_footage":95,  "on":0, "yield_per_ac":1.15, "pace":0.75, "home_price":385000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":2, "fence_cost_per_ff":94},
        {"front_footage":100, "on":0, "yield_per_ac":1.0,  "pace":0.75, "home_price":410000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":2, "fence_cost_per_ff":94},
    ]
    return {
        "project_name": name,
        "address": "",
        "gross_acreage": 0,
        "land_escalator": 0.05,
        "purchase_price_per_acre": 0,
        "closing_costs_pct": 0.045,
        "closing_date": "",
        "default_other_pct": 0.17,
        "sectional_other_pct": 0.17,       # Excel B6 = 0.17
        "landscaping_other_pct": 0.12,
        "contingency": 0.05,
        "site_work_pct": 0.01,
        "fenced_pct": 0.25,
        "cost_per_mailbox": 200,
        "cost_per_streetlight": 1700,
        "default_start_month": 1,
        "det_storage_rate": 1.1,            # Excel B31 = 1.1
        "det_depth": 9,                     # Excel B33 = 9
        "det_num_projects": 6,              # Excel B34 = 6
        "det_cost_per_cy": 10.0,            # Excel A37 = $10/CY
        "parks_pct": 0.03,                  # Excel B51 = 3%
        "drill_site_acres": 0,
        "commercial_pod_acres": 0,
        "residential_pod_acres": 0,
        "plants": [{"type":"None","notes":""} for _ in range(8)],
        "amenities": [{"type":"None","acres":0,"notes":""} for _ in range(6)],
        "other_netouts": [{"desc":"","acres":0,"notes":""} for _ in range(6)],
        "roads": [{"type":"","lf":0,"width":0,"road_setback":0,"landscaping_setback":0,"notes":""} for _ in range(6)],
        "takedowns": [{"period":0,"pct":0.5},{"period":36,"pct":0.5},{"period":0,"pct":0.0}],
        "plant_costs": [{"base_cost":0,"other_pct":0.17,"start_month":1,"ph2_base_cost":0,"ph2_other_pct":0.17,"ph2_start_month":37} for _ in range(8)],
        "amenity_costs": [{"base_cost":0,"other_pct":0.17,"start_month":1} for _ in range(6)],
        "det_costs": [{"other_pct":0.17,"landscaping_per_foot":2} for _ in range(6)],
        "other_costs": [{"base_cost":0,"other_pct":0.17,"start_month":1,"duration":1} for _ in range(6)],
        "road_costs": [{"other_pct":0.17,"start_month":1,"landscaping_per_sf":2,"light_spacing":0} for _ in range(6)],
        "lot_sizes": lot_size_defaults,
        "timing_method": "50/25/25",        # Excel B2 = 50/25/25
        "bem_period": 9,                    # Excel B3 = 9
        "bem_pct": 0.18,                    # Excel B4 = 18%
        "brokerage_fees": 0.03,             # Excel B5 = 3%
        "lot_closing_costs": 0.015,         # Excel B6 = 1.5%
        "take1_pct": 0.50,
        "take2_pct": 0.25,
        "take3_pct": 0.25,
        "price_per_ff": {str(yr): 1800 for yr in range(11)},
        "res_pod_acreage": 0,
        "res_pod_count": 1,
        "res_pods": [{"price_per_acre":120000,"closing_costs_pct":0.045,"implied_lots_per_acre":3.5,"impact_fee_per_lot":10000,"sale_period":12} for _ in range(6)],
        "comm_pod_acreage": 0,
        "comm_pod_count": 6,
        "comm_pods": [{"price_per_sf":8,"closing_costs_pct":0.045,"sale_period":12+i*24,"av_per_acre":1200000,"av_delay_months":18} for i in range(6)],
        "mud_bond": {"toggle":1,"amount":0,"reimbursement_pct":0.85,"first_bond_period":48,"bond_interval":12,"pct_to_dev":0.85,"receivables_fee":0.025,"debt_ratio":0.12},
        "wcid_bond": {"toggle":1,"amount":0,"reimbursement_pct":0.85,"first_bond_period":48,"bond_interval":12,"pct_to_dev":0.85,"receivables_fee":0.025,"debt_ratio":0.042},
        "marketing_pct": 0.02,
        "prof_svc_pct": 0.015,              # Excel B95 = 1.5%
        "dmf_pct": 0.025,                   # Excel B99 = 2.5%
        "personnel_monthly": 50000,         # Excel C103 = 50,000
        "marketing_personnel_monthly": 15000, # Excel C104 = 15,000
        "legal_monthly": 10000,             # Excel C108 = 10,000
        "mud_monthly": 35000,               # Excel C112 = 35,000
        "mud_pct": 0.2,                     # Excel D112 = 20% (what % of project MUD runs)
        "insurance_monthly": 10000,         # Excel C116 = 10,000
        "bookkeeping_monthly": 10000,       # Excel C120 = 10,000
    }

# ─── CHANGE LOG HELPERS ───────────────────────────────────────────────────────
_CHANGE_LOG_FIELDS = [
    ("purchase_price_per_acre", "Purchase Price / Acre"),
    ("gross_acreage",           "Gross Acreage"),
    ("land_escalator",          "Land Escalator"),
    ("contingency",             "Contingency"),
    ("closing_costs_pct",       "Closing Costs %"),
    ("net_dev_acres",           "Net Dev Acres"),
    ("dmf_rate",                "DMF Rate"),
    ("personnel_cost",          "Personnel Cost"),
]
_CHANGE_LOG_LOT_FIELDS = [
    ("home_price",   "Home Price"),
    ("pace",         "Pace (lots/mo)"),
    ("yield_per_ac", "Yield / Acre"),
    ("pct_mix",      "Mix %"),
]

def _compare_inputs(old_inp, new_inp):
    """Return list of {field, label, old, new} dicts for tracked changes."""
    changes = []
    for key, label in _CHANGE_LOG_FIELDS:
        ov, nv = old_inp.get(key), new_inp.get(key)
        if ov != nv and not (ov is None and nv is None):
            changes.append({"field": key, "label": label,
                            "old": ov, "new": nv})
    # Per-lot changes for active lots
    old_lots = old_inp.get("lot_sizes", [])
    new_lots = new_inp.get("lot_sizes", [])
    for i, new_lot in enumerate(new_lots):
        if not new_lot.get("on"):
            continue
        old_lot = old_lots[i] if i < len(old_lots) else {}
        ff = new_lot.get("ff", (i+1)*5+20)
        for key, label in _CHANGE_LOG_LOT_FIELDS:
            ov, nv = old_lot.get(key), new_lot.get(key)
            if ov != nv and not (ov is None and nv is None):
                changes.append({"field": f"lot_sizes[{i}].{key}",
                                "label": f"{label} ({ff}' lot)",
                                "old": ov, "new": nv})
    return changes


def _apply_sensitivity_override(inp, field, value):
    """Deep-copy inp, apply sensitivity override, return modified copy."""
    import copy
    inp2 = copy.deepcopy(inp)
    if field == "price_per_ff_base":
        # Scale all per-FF price years proportionally to the new base (year-0) value
        ppff = inp2.get("price_per_ff", {})
        ref = float(ppff.get("0", ppff.get(0, 1800)) or 1800)
        scale = float(value) / ref if ref else 1.0
        inp2["price_per_ff"] = {k: float(v or 0) * scale for k, v in ppff.items()}
    elif field == "lot_sizes.dev_cost_per_lot":
        # Scale wsd_per_ff and paving_per_ff proportionally across all active lots
        active = [r for r in inp2.get("lot_sizes", []) if r.get("on")]
        if active:
            costs = [(r.get("wsd_per_ff", 0) + r.get("paving_per_ff", 0)) * r.get("ff", 0)
                     for r in active]
            ref_avg = sum(costs) / len(costs) if costs else 0
            scale = float(value) / ref_avg if ref_avg else 1.0
            for row in inp2.get("lot_sizes", []):
                if row.get("on"):
                    row["wsd_per_ff"] = row.get("wsd_per_ff", 0) * scale
                    row["paving_per_ff"] = row.get("paving_per_ff", 0) * scale
    elif field.startswith("lot_sizes."):
        sub = field[len("lot_sizes."):]
        for row in inp2.get("lot_sizes", []):
            if row.get("on"):
                row[sub] = value
    else:
        inp2[field] = value
    return inp2


@app.route("/portfolio")
@login_required
def portfolio_page():
    if not session.get("is_admin"):
        return redirect(url_for("home"))
    pa = {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True, "portfolio": True}
    return render_template("portfolio.html", username=session.get("username"),
                           is_admin=True, page_access=pa)

@app.route("/api/portfolio", methods=["GET"])
@login_required
def portfolio():
    conn = get_db()
    cur = conn.cursor()
    include_archived = request.args.get("include_archived") == "true"
    where = "" if include_archived else "WHERE p.archived = FALSE"
    cur.execute(f"""
        SELECT p.id, p.name, p.address, p.outputs, COALESCE(p.status, 'Active') as status
        FROM projects p
        {where}
        ORDER BY p.name
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    result = []
    for r in rows:
        o = r["outputs"] or {}
        result.append({"id": r["id"], "name": r["name"], "address": r["address"],
                       "status": r["status"], "outputs": o})
    return jsonify(result)

@app.route("/api/projects/<int:pid>/export_excel", methods=["GET"])
@login_required
def export_excel(pid):
    try:
        from excel_export import export_excel as _export
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM projects WHERE id=%s", (pid,))
        proj = cur.fetchone()
        cur.close(); conn.close()
        if not proj:
            return jsonify({"error": "not found"}), 404
        inputs = proj["inputs"] or {}
        excel_bytes = _export(inputs)
        name = (proj.get("name") or "project").replace(" ", "_")
        return send_file(
            io.BytesIO(excel_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{name}_Underwriting.xlsx"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/projects/<int:pid>/backup", methods=["GET"])
@login_required
def backup_project(pid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM projects WHERE id=%s", (pid,))
    proj = dict(cur.fetchone() or {})
    cur.close(); conn.close()
    if not proj:
        return jsonify({"error": "not found"}), 404
    for k, v in proj.items():
        if hasattr(v, 'isoformat'):
            proj[k] = v.isoformat()
    name = (proj.get("name") or "project").replace(" ", "_")
    backup_data = json.dumps(proj, indent=2)
    return send_file(
        io.BytesIO(backup_data.encode()),
        mimetype="application/json",
        as_attachment=True,
        download_name=f"{name}_backup.json"
    )

@app.route("/api/projects/restore", methods=["POST"])
@login_required
def restore_project():
    data = request.json or {}
    inputs  = data.get("inputs", {})
    outputs = data.get("outputs", {})
    name    = data.get("name", "Restored Project")
    address = data.get("address", "")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO projects (name, address, created_by, inputs, outputs) VALUES (%s,%s,%s,%s,%s) RETURNING id",
        (name, address, session["user_id"], json.dumps(inputs), json.dumps(outputs))
    )
    new_id = cur.fetchone()["id"]
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "id": new_id})

@app.route("/api/projects/import_excel", methods=["POST"])
@login_required
def import_excel_project():
    """Upload an Ember underwriting Excel and create a new project from it."""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400
    try:
        from excel_import import import_excel
        file_bytes = f.read()
        inputs = import_excel(file_bytes)
    except Exception as e:
        return jsonify({"error": f"Failed to parse Excel: {e}"}), 400
    try:
        outputs = calculate(inputs)
    except Exception:
        outputs = {}
    name = inputs.get("project_name", "Imported Project")
    address = inputs.get("address", "")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO projects (name, address, created_by, inputs, outputs) VALUES (%s, %s, %s, %s, %s) RETURNING id",
        (name, address, session["user_id"], json.dumps(inputs), json.dumps(outputs))
    )
    pid = cur.fetchone()["id"]
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "id": pid, "name": name})

@app.route("/api/parse_excel", methods=["POST"])
@login_required
def parse_excel():
    """Parse an Ember underwriting Excel and return the inputs dict (no project created)."""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400
    try:
        from excel_import import import_excel
        file_bytes = f.read()
        inputs = import_excel(file_bytes)
    except Exception as e:
        return jsonify({"error": f"Failed to parse Excel: {e}"}), 400
    return jsonify({"ok": True, "inputs": inputs})

# ─── DASHBOARD REPORTS ────────────────────────────────────────────────────────
@app.route("/api/upload-dashboard", methods=["POST"])
@login_required
@admin_required
def upload_dashboard():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400
    try:
        file_bytes = f.read()
        data = parse_dashboard(file_bytes)
    except Exception as e:
        return jsonify({"error": f"Failed to parse file: {e}"}), 400
    conn = get_db()
    cur = conn.cursor()
    # Upsert returns
    cur.execute("DELETE FROM reports WHERE report_type = 'returns'")
    cur.execute(
        "INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
        ("returns", json.dumps(data.get("returns", {})), session["user_id"])
    )
    # Upsert loans
    cur.execute("DELETE FROM reports WHERE report_type = 'loans'")
    cur.execute(
        "INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
        ("loans", json.dumps(data.get("loans", {})), session["user_id"])
    )
    # Upsert operations
    cur.execute("DELETE FROM reports WHERE report_type = 'operations'")
    if data.get("operations"):
        cur.execute(
            "INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
            ("operations", json.dumps(data["operations"]), session["user_id"])
        )
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/export-returns-excel")
@login_required
def export_returns_excel():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("returns", True):
        return jsonify({"error": "Access denied"}), 403
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT data, uploaded_at FROM reports WHERE report_type = 'returns' ORDER BY uploaded_at DESC LIMIT 1")
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row or not row["data"]:
        return jsonify({"error": "No data available"}), 404

    data = row["data"]
    uploaded_at = row["uploaded_at"].strftime("%B %d, %Y") if row["uploaded_at"] else ""

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    LABEL_MAP = {"LP IRR": "Net Cashflow", "LP Equity Multiple": "Cumulative Net Cashflow"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Returns"

    # Light-background palette — legible on white Excel
    PROJ_FILL   = PatternFill("solid", fgColor="F2EFE8")   # warm tan for project header
    SUMM_FILL   = PatternFill("solid", fgColor="E8F0EE")   # light teal for summary header
    HEADER_FILL = PatternFill("solid", fgColor="F7F6F3")   # near-white for column headers
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    TEXT       = "1A1A1A"   # near-black for data
    HDR_TEXT   = "555555"   # medium grey for column header labels
    PROJ_TEXT  = "6B4E1E"   # dark brown for project title
    SUMM_TEXT  = "2D6B5A"   # dark teal for summary title
    ACCENT     = "7A5C1E"   # dark gold for highlighted metric rows (IRR/EM)

    def _f(bold=False, color=TEXT, size=9):
        return Font(name="Calibri", size=size, bold=bold, color=color)

    def _set_num(cell, val):
        """Write val as #,##0; blank if zero/None."""
        if isinstance(val, (int, float)) and val != 0:
            cell.value = val
            cell.number_format = "#,##0"
        else:
            cell.value = None

    years = data.get("years", [])
    all_idxs = list(range(len(years)))   # include every year column
    num_cols = 2 + len(years)            # label + Total + one per year

    r = 1
    ws.cell(row=r, column=1, value="Consolidated Ember Project Returns").font = Font(name="Calibri", bold=True, size=14, color=PROJ_TEXT)
    r += 1
    ws.cell(row=r, column=1, value=f"Last updated: {uploaded_at}  |  ($ in 000s)").font = _f(color="888888")
    r += 2

    # ── Project Returns Summary Table ──
    SUMMARY_HDR_FILL = PatternFill("solid", fgColor="EDE8DF")
    summary_cols = ["Project", "LP IRR", "Equity Multiple", "Total LP Profit", "Promote"]
    for ci, h in enumerate(summary_cols, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = _f(bold=True, color=HDR_TEXT)
        c.fill = SUMMARY_HDR_FILL
        c.border = cell_border
        c.alignment = Alignment(horizontal="left" if ci == 1 else "center")
    r += 1

    for proj in data.get("projects", []):
        metrics_by_label = {m["label"]: m for m in proj.get("metrics", [])}
        irr_val  = metrics_by_label.get("LP IRR", {}).get("total", None)
        em_val   = metrics_by_label.get("LP Equity Multiple", {}).get("total", None)
        pft_val  = metrics_by_label.get("Total LP Profit", {}).get("total", None)
        prom_val = metrics_by_label.get("Promote", {}).get("total", None)

        # Project name
        nc = ws.cell(row=r, column=1, value=proj["name"])
        nc.font = _f(bold=True, color=PROJ_TEXT)
        nc.border = cell_border

        # LP IRR — display as percentage
        ic = ws.cell(row=r, column=2)
        ic.font = _f(bold=True, color=ACCENT)
        ic.alignment = Alignment(horizontal="right")
        ic.border = cell_border
        if isinstance(irr_val, (int, float)) and irr_val:
            ic.value = irr_val
            ic.number_format = "0.0%"

        # Equity Multiple — display as multiplier
        ec = ws.cell(row=r, column=3)
        ec.font = _f(bold=True, color=ACCENT)
        ec.alignment = Alignment(horizontal="right")
        ec.border = cell_border
        if isinstance(em_val, (int, float)) and em_val:
            ec.value = em_val
            ec.number_format = '0.00"x"'

        # Total LP Profit
        pc = ws.cell(row=r, column=4)
        pc.font = _f()
        pc.alignment = Alignment(horizontal="right")
        pc.border = cell_border
        _set_num(pc, pft_val)

        # Promote
        prc = ws.cell(row=r, column=5)
        prc.font = _f()
        prc.alignment = Alignment(horizontal="right")
        prc.border = cell_border
        _set_num(prc, prom_val)

        r += 1

    r += 1  # blank spacer before detail sections

    def write_section_header(r, title, fill, color):
        c = ws.cell(row=r, column=1, value=title)
        c.font = Font(name="Calibri", bold=True, size=10, color=color)
        c.fill = fill
        c.border = cell_border
        for ci in range(2, num_cols + 1):
            cell = ws.cell(row=r, column=ci)
            cell.fill = fill
            cell.border = cell_border
        return r + 1

    def write_col_headers(r, col_labels):
        ws.cell(row=r, column=1, value="Metric").font = _f(bold=True, color=HDR_TEXT)
        ws.cell(row=r, column=1).fill = HEADER_FILL
        ws.cell(row=r, column=1).border = cell_border
        ws.cell(row=r, column=2, value="Total").font = _f(bold=True, color=HDR_TEXT)
        ws.cell(row=r, column=2).fill = HEADER_FILL
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2).border = cell_border
        for ci, lbl in enumerate(col_labels, 3):
            c = ws.cell(row=r, column=ci, value=lbl)
            c.font = _f(bold=True, color=HDR_TEXT)
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
            c.border = cell_border
        return r + 1

    def write_project(r, proj):
        metrics = proj.get("metrics", [])
        r = write_section_header(r, proj["name"], PROJ_FILL, PROJ_TEXT)
        r = write_col_headers(r, years)

        for m in metrics:
            label = m["label"]
            display = LABEL_MAP.get(label, label)
            is_accent = label in ("LP IRR", "LP Equity Multiple")
            txt_color = ACCENT if is_accent else TEXT

            # Total: for renamed rows use sum of yearly (Net Cashflow) or last non-zero (Cum. Net CF)
            if label == "LP IRR":
                total = sum(v for v in m.get("yearly", []) if isinstance(v, (int, float)))
            elif label == "LP Equity Multiple":
                yvals = [v for v in m.get("yearly", []) if isinstance(v, (int, float)) and v != 0]
                total = yvals[-1] if yvals else 0
            else:
                total = m.get("total", 0)

            lc = ws.cell(row=r, column=1, value=display)
            lc.font = _f(bold=is_accent, color=txt_color)
            lc.border = cell_border

            tc = ws.cell(row=r, column=2)
            tc.font = _f(bold=is_accent, color=txt_color)
            tc.alignment = Alignment(horizontal="right")
            tc.border = cell_border
            _set_num(tc, total)

            for ci, i in enumerate(all_idxs, 3):
                yc = ws.cell(row=r, column=ci)
                val = m["yearly"][i] if i < len(m.get("yearly", [])) else 0
                yc.font = _f(color=txt_color)
                yc.alignment = Alignment(horizontal="right")
                yc.border = cell_border
                _set_num(yc, val)
            r += 1
        return r + 1

    for proj in data.get("projects", []):
        r = write_project(r, proj)

    # Portfolio Summary
    summary = data.get("summary", [])
    if summary:
        r = write_section_header(r, "Portfolio Summary", SUMM_FILL, SUMM_TEXT)
        r = write_col_headers(r, years)
        for s in summary:
            lc = ws.cell(row=r, column=1, value=s["label"])
            lc.font = _f()
            lc.border = cell_border
            tc = ws.cell(row=r, column=2)
            tc.font = _f()
            tc.alignment = Alignment(horizontal="right")
            tc.border = cell_border
            _set_num(tc, s.get("total", 0))
            for ci, i in enumerate(all_idxs, 3):
                yc = ws.cell(row=r, column=ci)
                val = s["yearly"][i] if i < len(s.get("yearly", [])) else 0
                yc.font = _f()
                yc.alignment = Alignment(horizontal="right")
                yc.border = cell_border
                _set_num(yc, val)
            r += 1

    # Column widths — B=Total/IRR, C=EquityMult/yr1, D onwards = years
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 14   # "Equity Multiple" header needs a touch more
    ws.column_dimensions["D"].width = 14   # "Total LP Profit"
    ws.column_dimensions["E"].width = 13   # "Promote"
    for ci in range(6, 3 + len(years)):
        ws.column_dimensions[get_column_letter(ci)].width = 11

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from flask import send_file
    return send_file(output, as_attachment=True,
                     download_name="Ember_Project_Returns.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/returns")
@login_required
def returns_report():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("returns", True):
        return redirect(url_for("home"))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT data, uploaded_at FROM reports WHERE report_type = 'returns' ORDER BY uploaded_at DESC LIMIT 1")
    row = cur.fetchone()
    cur.close(); conn.close()
    data = row["data"] if row else None
    uploaded_at = row["uploaded_at"].strftime("%B %d, %Y") if row else None
    pa = session.get("page_access") or {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    if session.get("is_admin"):
        pa = {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    return render_template("returns.html", data=data, uploaded_at=uploaded_at, is_admin=session.get("is_admin"), page_access=pa)

@app.route("/loans")
@login_required
def loans_report():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("loans", True):
        return redirect(url_for("home"))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT data, uploaded_at FROM reports WHERE report_type = 'loans' ORDER BY uploaded_at DESC LIMIT 1")
    row = cur.fetchone()
    cur.close(); conn.close()
    data = row["data"] if row else None
    uploaded_at = row["uploaded_at"].strftime("%B %d, %Y") if row else None
    pa = session.get("page_access") or {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    if session.get("is_admin"):
        pa = {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    return render_template("loans.html", data=data, uploaded_at=uploaded_at, is_admin=session.get("is_admin"), page_access=pa)

@app.route("/api/export-operations-excel")
@login_required
def export_operations_excel():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("operations", True):
        return jsonify({"error": "Access denied"}), 403
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT data, uploaded_at FROM reports WHERE report_type = 'operations' ORDER BY uploaded_at DESC LIMIT 1")
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row or not row["data"]:
        return jsonify({"error": "No data available"}), 404

    data = row["data"]
    uploaded_at = row["uploaded_at"].strftime("%B %d, %Y") if row["uploaded_at"] else ""

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Operating Revenues"

    GOLD = "C8A96E"
    HEADER_FILL = PatternFill("solid", fgColor="1E2535")
    TOTALS_FILL = PatternFill("solid", fgColor="161B24")
    thin = Side(style="thin", color="2E3750")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr_font(bold=False):
        return Font(name="Calibri", size=9, bold=bold, color="8B95A8")

    def _val_font(bold=False):
        return Font(name="Calibri", size=9, bold=bold)

    def _gold_font(bold=True, size=10):
        return Font(name="Calibri", size=size, bold=bold, color=GOLD)

    def write_section(r, title):
        c = ws.cell(row=r, column=1, value=title)
        c.font = _gold_font(size=11)
        return r + 1

    def write_table(r, col_headers, data_rows, totals):
        # Header row
        for ci, h in enumerate(col_headers, 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = _hdr_font(bold=True)
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left")
            c.border = cell_border
        r += 1
        # Data rows
        for dr in data_rows:
            for ci, v in enumerate(dr, 1):
                cell = ws.cell(row=r, column=ci, value=v if v != 0 else None)
                cell.font = _val_font()
                cell.border = cell_border
                cell.alignment = Alignment(horizontal="left" if ci == 1 else "right")
                if ci > 1 and isinstance(v, (int, float)) and v:
                    cell.number_format = "#,##0"
            r += 1
        # Totals row
        ws.cell(row=r, column=1, value="Total").font = _val_font(bold=True)
        ws.cell(row=r, column=1).border = cell_border
        ws.cell(row=r, column=1).fill = TOTALS_FILL
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
        for ci, v in enumerate(totals, 2):
            cell = ws.cell(row=r, column=ci, value=v if v else None)
            cell.font = _val_font(bold=True)
            cell.fill = TOTALS_FILL
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="right")
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0"
        return r + 2

    r = 1
    # Title
    ws.cell(row=r, column=1, value="Ember Operating Revenues").font = Font(name="Calibri", bold=True, size=14, color=GOLD)
    r += 1
    ws.cell(row=r, column=1, value=f"Last updated: {uploaded_at}").font = Font(name="Calibri", size=9, color="8B95A8")
    r += 2

    # KPIs
    r = write_section(r, "KPI Summary")
    for kpi in data.get("kpis", []):
        ws.cell(row=r, column=1, value=kpi["label"]).font = _val_font()
        vc = ws.cell(row=r, column=2, value=kpi["value"])
        vc.font = _val_font(bold=True)
        vc.number_format = "#,##0"
        vc.alignment = Alignment(horizontal="right")
        r += 1
    r += 1

    # Annual Forecast
    yr = data.get("yearly_rollup", {})
    if yr.get("years"):
        r = write_section(r, "Annual Revenue Forecast (Next 5 Years)")
        headers = ["Revenue Source"] + [str(y) for y in yr["years"]]
        rows = [[row["label"]] + row["values"] for row in yr.get("rows", [])]
        r = write_table(r, headers, rows, yr.get("totals", []))

    # Monthly Revenue
    mo = data.get("monthly", {})
    if mo.get("dates"):
        r = write_section(r, "Monthly Fee Revenue")
        dates = mo["dates"]
        headers = ["Project / Category"] + [f"{d[5:7]}/{d[2:4]}" for d in dates]
        rows = [[f"{row['project']} — {row['category']}"] + row["values"] for row in mo.get("rows", [])]
        r = write_table(r, headers, rows, mo.get("totals", []))

    # Next 12 Months
    n12 = data.get("next_12_months", {})
    if n12.get("dates"):
        r = write_section(r, "Next 12 Months")
        dates = n12["dates"]
        headers = ["Revenue Source"] + [f"{d[5:7]}/{d[2:4]}" for d in dates]
        rows = [[row["label"]] + row["values"] for row in n12.get("rows", [])]
        r = write_table(r, headers, rows, n12.get("totals", []))

    # Next 12 Quarters
    qr = data.get("quarterly_rollup", {})
    if qr.get("quarters"):
        r = write_section(r, "Next 12 Quarters")
        headers = ["Revenue Source"] + qr["quarters"]
        rows = [[row["label"]] + row["values"] for row in qr.get("rows", [])]
        r = write_table(r, headers, rows, qr.get("totals", []))

    ws.column_dimensions["A"].width = 36
    for ci in range(2, 50):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(ci)].width = 11

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from flask import send_file
    return send_file(output, as_attachment=True,
                     download_name="Ember_Operating_Revenues.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/operations")
@login_required
def operations_report():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("operations", True):
        return redirect(url_for("home"))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT data, uploaded_at FROM reports WHERE report_type = 'operations' ORDER BY uploaded_at DESC LIMIT 1")
    row = cur.fetchone()
    cur.close(); conn.close()
    data = row["data"] if row else None
    uploaded_at = row["uploaded_at"].strftime("%B %d, %Y") if row else None
    pa = session.get("page_access") or {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    if session.get("is_admin"):
        pa = {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    return render_template("operations.html", data=data, uploaded_at=uploaded_at, is_admin=session.get("is_admin"), page_access=pa)

# ─── REPORT GENERATORS ────────────────────────────────────────────────────────

def _gen_excel_loans(data):
    """Generate loans Excel workbook bytes from report data."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loan Capacities"

    NAVY = "1A3A5C"
    GOLD = "C8A96E"
    HDR_FILL = PatternFill("solid", fgColor="1E2535")
    ALT_FILL = PatternFill("solid", fgColor="F7F8FA")
    thin = Side(style="thin", color="D0D5DD")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hf(bold=False, size=9): return Font(name="Calibri", size=size, bold=bold, color="8B95A8")
    def vf(bold=False, color="1A1A1A"): return Font(name="Calibri", size=9, bold=bold, color=color)
    def gf(size=12): return Font(name="Calibri", size=size, bold=True, color=GOLD)

    r = 1
    ws.cell(row=r, column=1, value="Loan Capacities & Debt Schedules").font = gf()
    r += 2

    def write_table(title, headers, rows, totals=None):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        r += 1
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = hf(bold=True)
            c.fill = HDR_FILL
            c.border = bdr
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left")
        r += 1
        for ri, row_data in enumerate(rows):
            fill = ALT_FILL if ri % 2 == 0 else PatternFill()
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font = vf()
                c.fill = fill
                c.border = bdr
                c.alignment = Alignment(horizontal="left" if ci == 1 else "right")
                if ci > 1 and isinstance(val, (int, float)):
                    c.number_format = "#,##0"
            r += 1
        if totals:
            tot_fill = PatternFill("solid", fgColor="E8EEF5")
            ws.cell(row=r, column=1, value="Total").font = vf(bold=True, color=NAVY)
            ws.cell(row=r, column=1).fill = tot_fill
            ws.cell(row=r, column=1).border = bdr
            for ci, val in enumerate(totals, 2):
                c = ws.cell(row=r, column=ci, value=val if val else None)
                c.font = vf(bold=True, color=NAVY)
                c.fill = tot_fill
                c.border = bdr
                c.alignment = Alignment(horizontal="right")
                if isinstance(val, (int, float)):
                    c.number_format = "#,##0"
            r += 1
        r += 1

    # MPC Loans table
    mpc = data.get("mpc_loans", {})
    if mpc.get("headers") and mpc.get("rows"):
        totals_row = [mpc["totals"].get(h, "") for h in mpc["headers"][1:]] if mpc.get("totals") else None
        write_table("MPC Loan Capacities", mpc["headers"],
                    [[row_d.get(h, "") for h in mpc["headers"]] for row_d in mpc["rows"]], totals_row)

    # Vertical Loans table
    vl = data.get("vertical_loans", {})
    if vl.get("headers") and vl.get("rows"):
        totals_row = [vl["totals"].get(h, "") for h in vl["headers"][1:]] if vl.get("totals") else None
        write_table("Vertical Loan Capacities", vl["headers"],
                    [[row_d.get(h, "") for h in vl["headers"]] for row_d in vl["rows"]], totals_row)

    # Debt Schedules — one mini-table per project
    for sched in data.get("debt_schedules", []):
        proj_name = sched.get("project", "Project")
        months = sched.get("months", [])
        if not months:
            continue
        headers = [""] + [str(m) for m in months]
        rows_data = [
            ["Scheduled Payments"] + [v for v in sched.get("payments", [])],
            ["Cumulative Payments"] + [v for v in sched.get("cumulative_payments", [])],
            ["Lot Revenues"] + [v for v in sched.get("revenues", [])],
            ["Cumulative Revenues"] + [v for v in sched.get("cumulative_revenues", [])],
        ]
        write_table(f"Debt Schedule — {proj_name}", headers, rows_data)

    ws.column_dimensions["A"].width = 30
    for ci in range(2, 30):
        ws.column_dimensions[get_column_letter(ci)].width = 12

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def _gen_pdf_report(report_type, data):
    """Generate a simple PDF for the given report type. Returns bytes."""
    from fpdf import FPDF

    class PDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 14)
            titles = {
                "returns": "Active Project Returns",
                "loans": "Loan Capacities & Debt Schedules",
                "operations": "Ember Operating Revenues",
            }
            self.set_text_color(26, 58, 92)
            self.cell(0, 10, titles.get(report_type, "Ember Report"), ln=True)
            self.set_font("Helvetica", "", 8)
            self.set_text_color(120, 120, 120)
            self.cell(0, 5, f"Generated {datetime.datetime.now().strftime('%B %d, %Y')}", ln=True)
            self.ln(3)

        def footer(self):
            self.set_y(-15)
            self.set_font("Helvetica", "", 7)
            self.set_text_color(150, 150, 150)
            self.cell(0, 10, f"Page {self.page_no()}", align="C")

    pdf = PDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    def draw_section(title):
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(26, 58, 92)
        pdf.cell(0, 7, title, ln=True)
        pdf.set_text_color(30, 30, 30)

    def draw_table(headers, rows, col_widths=None):
        if not headers:
            return
        n = len(headers)
        usable = pdf.w - pdf.l_margin - pdf.r_margin
        if col_widths is None:
            first = min(70, usable * 0.35)
            rest = (usable - first) / max(n - 1, 1)
            col_widths = [first] + [rest] * (n - 1)

        # Header row
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(30, 37, 53)
        pdf.set_text_color(139, 149, 168)
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], 5, str(h)[:30], border=0, fill=True,
                     align="L" if i == 0 else "R")
        pdf.ln()

        # Data rows
        pdf.set_font("Helvetica", "", 7)
        for ri, row_data in enumerate(rows):
            if pdf.get_y() > pdf.h - 25:
                pdf.add_page()
            fill = ri % 2 == 0
            pdf.set_fill_color(247, 248, 250) if fill else pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(30, 30, 30)
            for i, val in enumerate(row_data[:n]):
                txt = ""
                if isinstance(val, float):
                    txt = f"{val:,.0f}" if abs(val) >= 1 else f"{val:.1%}"
                elif isinstance(val, int):
                    txt = f"{val:,}"
                elif val is not None:
                    txt = str(val)[:30]
                pdf.cell(col_widths[i], 4.5, txt, border=0, fill=True,
                         align="L" if i == 0 else "R")
            pdf.ln()
        pdf.ln(3)

    if report_type == "returns":
        summary_cols = ["Project", "LP IRR", "Equity Multiple", "Total LP Profit", "Promote"]
        sum_rows = []
        for proj in data.get("projects", []):
            m = {m["label"]: m for m in proj.get("metrics", [])}
            sum_rows.append([
                proj["name"],
                f"{m.get('LP IRR',{}).get('total',0):.1%}" if m.get('LP IRR',{}).get('total') else "",
                f"{m.get('LP Equity Multiple',{}).get('total',0):.2f}x" if m.get('LP Equity Multiple',{}).get('total') else "",
                m.get('Total LP Profit',{}).get('total', ""),
                m.get('Promote',{}).get('total', ""),
            ])
        draw_section("Portfolio Summary")
        draw_table(summary_cols, sum_rows, [70, 22, 28, 30, 25])
        years = data.get("years", [])
        for proj in data.get("projects", []):
            draw_section(proj["name"])
            hdrs = ["Metric", "Total"] + [str(y) for y in years[:10]]
            rows_data = []
            for m in proj.get("metrics", []):
                row_vals = [m["label"], m.get("total", "")] + (m.get("yearly", [])[:10])
                rows_data.append(row_vals)
            draw_table(hdrs, rows_data)

    elif report_type == "loans":
        mpc = data.get("mpc_loans", {})
        if mpc.get("headers") and mpc.get("rows"):
            draw_section("MPC Loan Capacities")
            rows_data = [[r.get(h, "") for h in mpc["headers"]] for r in mpc["rows"]]
            draw_table(mpc["headers"], rows_data)
        vl = data.get("vertical_loans", {})
        if vl.get("headers") and vl.get("rows"):
            draw_section("Vertical Loan Capacities")
            rows_data = [[r.get(h, "") for h in vl["headers"]] for r in vl["rows"]]
            draw_table(vl["headers"], rows_data)
        for sched in data.get("debt_schedules", []):
            months = sched.get("months", [])
            if not months:
                continue
            draw_section(f"Debt Schedule — {sched.get('project','')}")
            hdrs = [""] + [str(m) for m in months[:12]]
            rows_data = [
                ["Scheduled Payments"] + sched.get("payments", [])[:12],
                ["Cumulative Payments"] + sched.get("cumulative_payments", [])[:12],
                ["Lot Revenues"] + sched.get("revenues", [])[:12],
                ["Cumulative Revenues"] + sched.get("cumulative_revenues", [])[:12],
            ]
            draw_table(hdrs, rows_data)

    elif report_type == "operations":
        for kpi in data.get("kpis", []):
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(30, 30, 30)
            pdf.cell(80, 5, kpi.get("label", ""), ln=False)
            pdf.set_font("Helvetica", "B", 9)
            val = kpi.get("value", "")
            txt = f"{val:,}" if isinstance(val, (int, float)) else str(val)
            pdf.cell(0, 5, txt, ln=True)
        pdf.ln(3)
        yr = data.get("yearly_rollup", {})
        if yr.get("years"):
            draw_section("Annual Revenue Forecast")
            hdrs = ["Revenue Source"] + [str(y) for y in yr["years"]]
            rows_data = [[row["label"]] + row.get("values", []) for row in yr.get("rows", [])]
            draw_table(hdrs, rows_data)
        mo = data.get("monthly", {})
        if mo.get("dates"):
            draw_section("Monthly Fee Revenue")
            dates = mo["dates"][:12]
            hdrs = ["Project / Category"] + [f"{d[5:7]}/{d[2:4]}" for d in dates]
            rows_data = [[f"{r['project']} — {r['category']}"] + r.get("values", [])[:12]
                         for r in mo.get("rows", [])]
            draw_table(hdrs, rows_data)

    return pdf.output()


def _send_monthly_emails(force=False):
    """Send monthly reports to all opted-in users. Returns count of emails sent."""
    now = datetime.datetime.utcnow()
    period = now.strftime("%Y-%m")

    conn = get_db()
    cur = conn.cursor()

    if not force:
        cur.execute("SELECT id FROM report_sends WHERE period = %s", (period,))
        if cur.fetchone():
            cur.close(); conn.close()
            return 0  # already sent this month

    # Fetch opted-in users with emails
    cur.execute("SELECT id, username, email, report_format FROM users WHERE report_opt_in = TRUE AND email IS NOT NULL AND email != ''")
    recipients = cur.fetchall()

    # Fetch latest report data for all three types
    report_data = {}
    for rt in ("returns", "loans", "operations"):
        cur.execute("SELECT data FROM reports WHERE report_type = %s ORDER BY uploaded_at DESC LIMIT 1", (rt,))
        row = cur.fetchone()
        report_data[rt] = row["data"] if row else None

    cur.close(); conn.close()

    if not recipients:
        return 0

    sendgrid_key = os.environ.get("SENDGRID_API_KEY", "")
    from_addr = os.environ.get("SMTP_FROM", "")

    if not sendgrid_key:
        raise ValueError("SENDGRID_API_KEY environment variable must be set")
    if not from_addr:
        raise ValueError("SMTP_FROM environment variable must be set (used as sender address)")

    subject = now.strftime("%B %Y") + " Ember Reports"

    report_labels = {
        "returns": "Active Project Returns",
        "loans": "Loan Capacities & Debt Schedules",
        "operations": "Ember Operating Revenues",
    }

    sg = SendGridAPIClient(sendgrid_key)
    sent_count = 0

    for user in recipients:
        fmt = user["report_format"] or "pdf"
        email_addr = user["email"]

        body_lines = [
            f"Hello {user['username']},",
            "",
            f"Please find your {now.strftime('%B %Y')} Ember reports attached below.",
            "",
            "The following reports are included:",
        ]
        for rt, label in report_labels.items():
            if report_data.get(rt):
                body_lines.append(f"  • {label} ({fmt.upper()} format)")
        body_lines += [
            "",
            "These reports are generated automatically on the 1st of each month.",
            "",
            "— Ember Acquisitions",
        ]

        message = Mail(
            from_email=from_addr,
            to_emails=email_addr,
            subject=subject,
            plain_text_content="\n".join(body_lines),
        )

        for rt, label in report_labels.items():
            if not report_data.get(rt):
                continue
            data = report_data[rt]
            try:
                if fmt == "excel":
                    if rt == "returns":
                        file_bytes = _gen_excel_returns(data)
                        filename = f"{label.replace(' ','_')}.xlsx"
                        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    elif rt == "loans":
                        file_bytes = _gen_excel_loans(data)
                        filename = "Loan_Capacities.xlsx"
                        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    else:
                        file_bytes = _gen_excel_operations(data)
                        filename = "Ember_Operating_Revenues.xlsx"
                        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                else:
                    file_bytes = bytes(_gen_pdf_report(rt, data))
                    filename = f"{label.replace(' ','_')}.pdf"
                    mime_type = "application/pdf"

                attachment = Attachment(
                    FileContent(base64.b64encode(file_bytes).decode()),
                    FileName(filename),
                    FileType(mime_type),
                    Disposition("attachment"),
                )
                message.attachment = attachment
            except Exception as e:
                print(f"Error generating {rt} {fmt}: {e}")

        try:
            sg.send(message)
            sent_count += 1
        except Exception as e:
            print(f"SendGrid error sending to {email_addr}: {e}")

    # Record successful send (skip if forced to allow re-testing)
    if not force:
        conn2 = get_db()
        cur2 = conn2.cursor()
        cur2.execute("INSERT INTO report_sends (period) VALUES (%s) ON CONFLICT DO NOTHING", (period,))
        conn2.commit(); cur2.close(); conn2.close()

    return sent_count


def _gen_excel_returns(data):
    """Extract the returns Excel generation logic for reuse in email sending."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    LABEL_MAP = {"LP IRR": "Net Cashflow", "LP Equity Multiple": "Cumulative Net Cashflow"}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Returns"

    PROJ_FILL   = PatternFill("solid", fgColor="F2EFE8")
    SUMM_FILL   = PatternFill("solid", fgColor="E8F0EE")
    HEADER_FILL = PatternFill("solid", fgColor="F7F6F3")
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    TEXT="1A1A1A"; HDR_TEXT="555555"; PROJ_TEXT="6B4E1E"; SUMM_TEXT="2D6B5A"; ACCENT="7A5C1E"

    def _f(bold=False, color=TEXT, size=9):
        return Font(name="Calibri", size=size, bold=bold, color=color)
    def _set_num(cell, val):
        if isinstance(val, (int, float)) and val != 0:
            cell.value = val; cell.number_format = "#,##0"
        else:
            cell.value = None

    years = data.get("years", [])
    all_idxs = list(range(len(years)))
    num_cols = 2 + len(years)
    r = 1
    ws.cell(row=r, column=1, value="Consolidated Ember Project Returns").font = Font(name="Calibri", bold=True, size=14, color=PROJ_TEXT)
    r += 1
    ws.cell(row=r, column=1, value="($ in 000s)").font = _f(color="888888")
    r += 2

    SUMMARY_HDR_FILL = PatternFill("solid", fgColor="EDE8DF")
    summary_cols = ["Project", "LP IRR", "Equity Multiple", "Total LP Profit", "Promote"]
    for ci, h in enumerate(summary_cols, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = _f(bold=True, color=HDR_TEXT); c.fill = SUMMARY_HDR_FILL; c.border = cell_border
        c.alignment = Alignment(horizontal="left" if ci==1 else "center")
    r += 1
    for proj in data.get("projects", []):
        metrics_by_label = {m["label"]: m for m in proj.get("metrics", [])}
        irr_val = metrics_by_label.get("LP IRR", {}).get("total"); em_val = metrics_by_label.get("LP Equity Multiple", {}).get("total")
        pft_val = metrics_by_label.get("Total LP Profit", {}).get("total"); prom_val = metrics_by_label.get("Promote", {}).get("total")
        nc = ws.cell(row=r, column=1, value=proj["name"]); nc.font = _f(bold=True, color=PROJ_TEXT); nc.border = cell_border
        ic = ws.cell(row=r, column=2); ic.font = _f(bold=True, color=ACCENT); ic.alignment = Alignment(horizontal="right"); ic.border = cell_border
        if isinstance(irr_val, (int, float)) and irr_val: ic.value = irr_val; ic.number_format = "0.0%"
        ec = ws.cell(row=r, column=3); ec.font = _f(bold=True, color=ACCENT); ec.alignment = Alignment(horizontal="right"); ec.border = cell_border
        if isinstance(em_val, (int, float)) and em_val: ec.value = em_val; ec.number_format = '0.00"x"'
        pc = ws.cell(row=r, column=4); pc.font = _f(); pc.alignment = Alignment(horizontal="right"); pc.border = cell_border; _set_num(pc, pft_val)
        prc = ws.cell(row=r, column=5); prc.font = _f(); prc.alignment = Alignment(horizontal="right"); prc.border = cell_border; _set_num(prc, prom_val)
        r += 1
    r += 1

    def write_section_header(r, title, fill, color):
        c = ws.cell(row=r, column=1, value=title); c.font = Font(name="Calibri", bold=True, size=10, color=color); c.fill = fill; c.border = cell_border
        for ci in range(2, num_cols+1): cell=ws.cell(row=r, column=ci); cell.fill=fill; cell.border=cell_border
        return r+1
    def write_col_headers(r, col_labels):
        ws.cell(row=r, column=1, value="Metric").font = _f(bold=True, color=HDR_TEXT); ws.cell(row=r, column=1).fill=HEADER_FILL; ws.cell(row=r, column=1).border=cell_border
        ws.cell(row=r, column=2, value="Total").font = _f(bold=True, color=HDR_TEXT); ws.cell(row=r, column=2).fill=HEADER_FILL; ws.cell(row=r, column=2).alignment=Alignment(horizontal="center"); ws.cell(row=r, column=2).border=cell_border
        for ci, lbl in enumerate(col_labels, 3):
            c=ws.cell(row=r, column=ci, value=lbl); c.font=_f(bold=True, color=HDR_TEXT); c.fill=HEADER_FILL; c.alignment=Alignment(horizontal="center"); c.border=cell_border
        return r+1
    def write_project(r, proj):
        r=write_section_header(r, proj["name"], PROJ_FILL, PROJ_TEXT); r=write_col_headers(r, years)
        for m in proj.get("metrics", []):
            label=m["label"]; display=LABEL_MAP.get(label, label); is_accent=label in ("LP IRR","LP Equity Multiple"); txt_color=ACCENT if is_accent else TEXT
            total = sum(v for v in m.get("yearly",[]) if isinstance(v,(int,float))) if label=="LP IRR" else ([v for v in m.get("yearly",[]) if isinstance(v,(int,float)) and v!=0] or [0])[-1] if label=="LP Equity Multiple" else m.get("total",0)
            lc=ws.cell(row=r, column=1, value=display); lc.font=_f(bold=is_accent, color=txt_color); lc.border=cell_border
            tc=ws.cell(row=r, column=2); tc.font=_f(bold=is_accent, color=txt_color); tc.alignment=Alignment(horizontal="right"); tc.border=cell_border; _set_num(tc, total)
            for ci, i in enumerate(all_idxs, 3):
                yc=ws.cell(row=r, column=ci); val=m["yearly"][i] if i<len(m.get("yearly",[])) else 0; yc.font=_f(color=txt_color); yc.alignment=Alignment(horizontal="right"); yc.border=cell_border; _set_num(yc, val)
            r+=1
        return r+1

    for proj in data.get("projects", []):
        r = write_project(r, proj)
    summary = data.get("summary", [])
    if summary:
        r=write_section_header(r, "Portfolio Summary", SUMM_FILL, SUMM_TEXT); r=write_col_headers(r, years)
        for s in summary:
            lc=ws.cell(row=r, column=1, value=s["label"]); lc.font=_f(); lc.border=cell_border
            tc=ws.cell(row=r, column=2); tc.font=_f(); tc.alignment=Alignment(horizontal="right"); tc.border=cell_border; _set_num(tc, s.get("total",0))
            for ci, i in enumerate(all_idxs, 3):
                yc=ws.cell(row=r, column=ci); val=s["yearly"][i] if i<len(s.get("yearly",[])) else 0; yc.font=_f(); yc.alignment=Alignment(horizontal="right"); yc.border=cell_border; _set_num(yc, val)
            r+=1

    ws.column_dimensions["A"].width=32; ws.column_dimensions["B"].width=13
    ws.column_dimensions["C"].width=14; ws.column_dimensions["D"].width=14; ws.column_dimensions["E"].width=13
    for ci in range(6, 3+len(years)): ws.column_dimensions[get_column_letter(ci)].width=11

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.read()


def _gen_excel_operations(data):
    """Extract the operations Excel generation logic for reuse in email sending."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Operating Revenues"
    GOLD="C8A96E"; HEADER_FILL=PatternFill("solid", fgColor="1E2535"); TOTALS_FILL=PatternFill("solid", fgColor="161B24")
    thin=Side(style="thin", color="2E3750"); cell_border=Border(left=thin, right=thin, top=thin, bottom=thin)
    def _hdr_font(bold=False): return Font(name="Calibri", size=9, bold=bold, color="8B95A8")
    def _val_font(bold=False): return Font(name="Calibri", size=9, bold=bold)
    def write_section(r, title):
        c=ws.cell(row=r, column=1, value=title); c.font=Font(name="Calibri", size=11, bold=True, color=GOLD); return r+1
    def write_table(r, col_headers, data_rows, totals):
        for ci, h in enumerate(col_headers, 1):
            c=ws.cell(row=r, column=ci, value=h); c.font=_hdr_font(bold=True); c.fill=HEADER_FILL; c.border=cell_border; c.alignment=Alignment(horizontal="left" if ci==1 else "center")
        r+=1
        for ri, row_data in enumerate(data_rows):
            for ci, val in enumerate(row_data, 1):
                c=ws.cell(row=r, column=ci, value=val if val else None); c.font=_val_font(); c.border=cell_border; c.alignment=Alignment(horizontal="left" if ci==1 else "right")
                if ci>1 and isinstance(val, (int,float)): c.number_format="#,##0"
            r+=1
        ws.cell(row=r, column=1, value="Total").font=_val_font(bold=True); ws.cell(row=r, column=1).border=cell_border; ws.cell(row=r, column=1).fill=TOTALS_FILL; ws.cell(row=r, column=1).alignment=Alignment(horizontal="left")
        for ci, v in enumerate(totals, 2):
            cell=ws.cell(row=r, column=ci, value=v if v else None); cell.font=_val_font(bold=True); cell.fill=TOTALS_FILL; cell.border=cell_border; cell.alignment=Alignment(horizontal="right")
            if isinstance(v, (int,float)): cell.number_format="#,##0"
        return r+2

    r=1; ws.cell(row=r, column=1, value="Ember Operating Revenues").font=Font(name="Calibri", bold=True, size=14, color=GOLD); r+=2
    r=write_section(r, "KPI Summary")
    for kpi in data.get("kpis", []):
        ws.cell(row=r, column=1, value=kpi["label"]).font=_val_font()
        vc=ws.cell(row=r, column=2, value=kpi["value"]); vc.font=_val_font(bold=True); vc.number_format="#,##0"; vc.alignment=Alignment(horizontal="right"); r+=1
    r+=1
    yr=data.get("yearly_rollup",{})
    if yr.get("years"):
        r=write_section(r, "Annual Revenue Forecast (Next 5 Years)")
        r=write_table(r, ["Revenue Source"]+[str(y) for y in yr["years"]], [[row["label"]]+row["values"] for row in yr.get("rows",[])], yr.get("totals",[]))
    mo=data.get("monthly",{})
    if mo.get("dates"):
        r=write_section(r, "Monthly Fee Revenue")
        r=write_table(r, ["Project / Category"]+[f"{d[5:7]}/{d[2:4]}" for d in mo["dates"]], [[f"{row['project']} — {row['category']}"]+row["values"] for row in mo.get("rows",[])], mo.get("totals",[]))
    n12=data.get("next_12_months",{})
    if n12.get("dates"):
        r=write_section(r, "Next 12 Months")
        r=write_table(r, ["Revenue Source"]+[f"{d[5:7]}/{d[2:4]}" for d in n12["dates"]], [[row["label"]]+row["values"] for row in n12.get("rows",[])], n12.get("totals",[]))
    qr=data.get("quarterly_rollup",{})
    if qr.get("quarters"):
        r=write_section(r, "Next 12 Quarters")
        r=write_table(r, ["Revenue Source"]+qr["quarters"], [[row["label"]]+row["values"] for row in qr.get("rows",[])], qr.get("totals",[]))

    ws.column_dimensions["A"].width=36
    for ci in range(2, 50): ws.column_dimensions[get_column_letter(ci)].width=11
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return out.read()


# ─── SCHEDULER ────────────────────────────────────────────────────────────────
def _start_scheduler():
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        scheduler = BackgroundScheduler()
        # Run on the 1st of every month at 8:00 AM UTC
        scheduler.add_job(_send_monthly_emails, "cron", day=1, hour=8, minute=0)
        scheduler.start()
        print("APScheduler started — monthly report job scheduled for 1st of each month at 08:00 UTC")
    except Exception as e:
        print(f"Scheduler failed to start: {e}")

_start_scheduler()


if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=5001)
